from __future__ import annotations

import re
from datetime import datetime
from typing import Any, Dict, List, Optional

from flask import Blueprint, jsonify, render_template, request
from openpyxl import load_workbook

from config import REPORT_BASE_DIR
from log_utils import log_activity


disallineamento_qr_bp = Blueprint("disallineamento_qr", __name__)

REPORT_DIR = REPORT_BASE_DIR
EXCEL_PATH = REPORT_DIR / "Incidenti_robot.xlsx"

ROBOT_IDS = [
    "16278", "16279", "16292", "16294", "16302", "16306",
    "16313", "16314", "16325", "16337", "16339", "16340", "16348", "16349", "16350",
]

EVENT_TYPE_QR = "qr"
EVENT_TYPE_FW = "fw"
EVENT_TYPE_MCU = "mcu"
EVENT_TYPE_BAT = "bat"
EVENT_TYPE_MAN_ORD = "man_ord"
EVENT_TYPE_MANSTRA_SW = "manstra_sw"
EVENT_TYPE_MANSTRA_HW = "manstra_hw"

EVENT_TYPE_ORDER = [
    EVENT_TYPE_QR,
    EVENT_TYPE_FW,
    EVENT_TYPE_MCU,
    EVENT_TYPE_BAT,
    EVENT_TYPE_MAN_ORD,
    EVENT_TYPE_MANSTRA_SW,
    EVENT_TYPE_MANSTRA_HW,
]

EVENT_LABEL_PREFIX = {
    EVENT_TYPE_QR: "QR",
    EVENT_TYPE_FW: "FW",
    EVENT_TYPE_MCU: "MCU",
    EVENT_TYPE_BAT: "BAT",
    EVENT_TYPE_MAN_ORD: "MAN",
    EVENT_TYPE_MANSTRA_SW: "STR",
    EVENT_TYPE_MANSTRA_HW: "STR",
}

PALETTE = [
    "#4da3ff",
    "#7cc4ff",
    "#8f7cff",
    "#62e6a5",
    "#f39c12",
    "#f08a5d",
    "#c678dd",
    "#ff6b6b",
]

VERSION_RE = re.compile(r"\bv?(\d+(?:\.\d+){1,3})\b", re.IGNORECASE)
CAMERA_RE = re.compile(r"\b(camera|telecamera|cam)\b", re.IGNORECASE)
MCU_RE = re.compile(r"\b(mcu|scheda madre|mainboard|motherboard|controller board|control board)\b", re.IGNORECASE)
SOFTWARE_RE = re.compile(r"\b(software|firmware|update|upgrade|versione|version)\b", re.IGNORECASE)
BATTERY_RE = re.compile(r"\b(batteria|battery)\b", re.IGNORECASE)
HARDWARE_HINT_RE = re.compile(
    r"\b(motore|motor|ruota|wheel|sensore|sensor|pedana|lifter|bumper|scheda|board|camera|telecamera|charger|caricatore|ricambio|sostit)\b",
    re.IGNORECASE,
)


def _normalize_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_key(value: Any) -> str:
    text = _normalize_str(value).lower()
    return re.sub(r"[\s_]+", "", text)


def _parse_date(date_str: str) -> Optional[datetime]:
    raw = _normalize_str(date_str)
    if not raw:
        return None

    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(raw, fmt)
        except Exception:
            continue
    return None


def _format_date_short(date_str: str) -> str:
    dt = _parse_date(date_str)
    if not dt:
        return _normalize_str(date_str)
    return dt.strftime("%d/%m/%y")


def _header_index_map(ws) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for idx, cell in enumerate(ws[1]):
        key = _normalize_key(cell.value)
        if key:
            out[key] = idx
    return out


def _get_cell(row: tuple, header_map: Dict[str, int], *keys: str) -> Any:
    for key in keys:
        idx = header_map.get(_normalize_key(key))
        if idx is None:
            continue
        if idx < len(row):
            return row[idx]
    return None


def _build_text_blob(*values: Any) -> str:
    return " ".join(_normalize_str(v).lower() for v in values if _normalize_str(v))


def _extract_robot_ids(robot_value: Any) -> List[str]:
    text = _normalize_str(robot_value)
    if not text:
        return []

    found: List[str] = []
    seen: set[str] = set()
    for robot_id in ROBOT_IDS:
        if robot_id in text and robot_id not in seen:
            found.append(robot_id)
            seen.add(robot_id)
    return found


def _split_parts(raw_value: Any) -> List[str]:
    raw = _normalize_str(raw_value)
    if not raw:
        return []

    cleaned = raw.replace("_x000D_", "\n")
    chunks = re.split(r"[\n;,|]+", cleaned)

    out: List[str] = []
    seen: set[str] = set()
    for chunk in chunks:
        text = " ".join(chunk.split()).strip(" -")
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        out.append(text)
        seen.add(key)
    return out


def _extract_versions(*values: Any) -> List[str]:
    blob = " ".join(_normalize_str(v) for v in values if _normalize_str(v))
    versions: List[str] = []
    seen: set[str] = set()
    for match in VERSION_RE.finditer(blob):
        version = match.group(1)
        if version not in seen:
            versions.append(version)
            seen.add(version)
    return versions[:4]


def _extract_components(text: str) -> List[str]:
    out: List[str] = []
    if MCU_RE.search(text):
        out.append("MCU")
    if CAMERA_RE.search(text):
        out.append("CAM")
    return out


def _extract_hardware_parts(row: Dict[str, Any]) -> List[str]:
    explicit = _split_parts(row.get("parti_coinvolte"))
    if explicit:
        return explicit[:3]

    candidates = [
        row.get("titolo"),
        row.get("risoluzione"),
        row.get("note"),
        row.get("errore"),
    ]
    out: List[str] = []
    seen: set[str] = set()
    for value in candidates:
        text = _normalize_str(value)
        if not text:
            continue
        if HARDWARE_HINT_RE.search(text):
            compact = re.sub(r"\s+", " ", text).strip()
            short = compact[:60] + ("…" if len(compact) > 60 else "")
            key = short.lower()
            if key not in seen:
                out.append(short)
                seen.add(key)
        if len(out) >= 3:
            break
    return out


def _version_badge_color(version: str) -> str:
    if not version:
        return "#5a6b7c"
    return PALETTE[sum(ord(ch) for ch in version) % len(PALETTE)]


def _categorize_event(row: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    categoria = _normalize_str(row.get("categoria")).lower()
    titolo = row.get("titolo")
    parti_coinvolte = row.get("parti_coinvolte")
    note = row.get("note")
    risoluzione = row.get("risoluzione")
    errore = row.get("errore")

    focus_text = _build_text_blob(titolo, parti_coinvolte, risoluzione)
    full_text = _build_text_blob(titolo, parti_coinvolte, note, risoluzione, errore)

    is_qr_misalignment = (
        ("dissalineato" in categoria or "disallineato" in categoria)
        and "qr" in categoria
    )
    if is_qr_misalignment:
        return {
            "event_type": EVENT_TYPE_QR,
            "detail_category": "Disallineamento QR",
            "software_components": [],
            "software_versions": [],
            "hardware_parts": [],
        }

    is_manutenzione_stra = "intervento manutenzione straordinaria" in categoria
    is_manutenzione = is_manutenzione_stra or "intervento manutenzione" in categoria
    if not is_manutenzione:
        return None

    if is_manutenzione_stra:
        software_components = _extract_components(full_text)
        software_versions = _extract_versions(titolo, parti_coinvolte, note, risoluzione)
        software_signal = bool(software_components or software_versions or SOFTWARE_RE.search(full_text) or CAMERA_RE.search(full_text))

        if software_signal:
            return {
                "event_type": EVENT_TYPE_MANSTRA_SW,
                "detail_category": "Manutenzione straordinaria software",
                "software_components": software_components,
                "software_versions": software_versions,
                "hardware_parts": [],
            }

        return {
            "event_type": EVENT_TYPE_MANSTRA_HW,
            "detail_category": "Manutenzione straordinaria hardware",
            "software_components": [],
            "software_versions": [],
            "hardware_parts": _extract_hardware_parts(row),
        }

    if BATTERY_RE.search(focus_text):
        return {
            "event_type": EVENT_TYPE_BAT,
            "detail_category": "Cambio batteria",
            "software_components": [],
            "software_versions": [],
            "hardware_parts": [],
        }

    if MCU_RE.search(focus_text):
        return {
            "event_type": EVENT_TYPE_MCU,
            "detail_category": "Cambio MCU / scheda madre",
            "software_components": ["MCU"],
            "software_versions": _extract_versions(titolo, parti_coinvolte, risoluzione),
            "hardware_parts": [],
        }

    if SOFTWARE_RE.search(focus_text) or CAMERA_RE.search(focus_text):
        return {
            "event_type": EVENT_TYPE_FW,
            "detail_category": "Intervento software / firmware",
            "software_components": _extract_components(full_text),
            "software_versions": _extract_versions(titolo, parti_coinvolte, risoluzione),
            "hardware_parts": [],
        }

    return {
        "event_type": EVENT_TYPE_MAN_ORD,
        "detail_category": "Manutenzione ordinaria",
        "software_components": [],
        "software_versions": [],
        "hardware_parts": _extract_hardware_parts(row),
    }


def _read_events_from_excel() -> List[Dict[str, Any]]:
    events: List[Dict[str, Any]] = []

    try:
        if not EXCEL_PATH.exists() or not EXCEL_PATH.is_file():
            return events

        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        header_map = _header_index_map(ws)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            date_str = _normalize_str(_get_cell(row, header_map, "data", "Data"))
            date_obj = _parse_date(date_str)
            if not date_obj:
                continue

            raw_row = {
                "id": _get_cell(row, header_map, "id", "ID"),
                "data": date_str,
                "ora": _normalize_str(_get_cell(row, header_map, "ora", "Ora")),
                "categoria": _normalize_str(_get_cell(row, header_map, "Categoria", "categoria")),
                "titolo": _normalize_str(_get_cell(row, header_map, "Titolo", "titolo")),
                "robot": _normalize_str(_get_cell(row, header_map, "robot", "Robot")),
                "note": _normalize_str(_get_cell(row, header_map, "note", "Note")),
                "errore": _normalize_str(_get_cell(row, header_map, "errore", "Errore")),
                "risoluzione": _normalize_str(_get_cell(row, header_map, "risoluzione", "Risoluzione")),
                "parti_coinvolte": _normalize_str(_get_cell(row, header_map, "parti_coinvolte", "Parti Coinvolte", "parti coinvolte")),
            }

            robots = _extract_robot_ids(raw_row["robot"])
            if not robots:
                continue

            classification = _categorize_event(raw_row)
            if not classification:
                continue

            try:
                report_id = int(str(raw_row.get("id")).strip())
            except Exception:
                report_id = None

            titolo = raw_row["titolo"] or raw_row["categoria"] or "Evento"

            detail_lines: List[str] = [classification["detail_category"]]
            if classification["software_components"]:
                detail_lines.append("Componenti SW: " + ", ".join(classification["software_components"]))
            if classification["software_versions"]:
                detail_lines.append("Versioni: " + " • ".join(classification["software_versions"]))
            if classification["hardware_parts"]:
                detail_lines.append("Parti: " + " • ".join(classification["hardware_parts"]))
            if raw_row["parti_coinvolte"] and not classification["hardware_parts"]:
                detail_lines.append("Parti: " + raw_row["parti_coinvolte"])
            if raw_row["risoluzione"]:
                detail_lines.append("Risoluzione: " + raw_row["risoluzione"])

            events.append(
                {
                    "report_id": report_id,
                    "report_url": f"/MedicairGeek/storicoReport/{report_id}" if report_id is not None else "",
                    "date": date_obj,
                    "date_str": date_str,
                    "date_short": _format_date_short(date_str),
                    "robots": robots,
                    "titolo": titolo,
                    "categoria": raw_row["categoria"],
                    "event_type": classification["event_type"],
                    "detail_category": classification["detail_category"],
                    "software_components": classification["software_components"],
                    "software_versions": classification["software_versions"],
                    "hardware_parts": classification["hardware_parts"],
                    "detail_text": "\n".join(detail_lines),
                }
            )

        wb.close()
    except Exception as exc:
        print(f"[DisallineamentoQR] Errore lettura Excel: {exc}")

    return events


def _build_robot_events_table() -> tuple[List[Dict[str, Any]], List[Dict[str, str]], Dict[str, Dict[str, int]]]:
    events = _read_events_from_excel()

    robot_events_map: Dict[str, List[Dict[str, Any]]] = {}
    all_dates: Dict[str, str] = {}

    for event in events:
        all_dates[event["date_str"]] = event["date_short"]
        for robot_id in event["robots"]:
            robot_events_map.setdefault(robot_id, []).append(dict(event))

    unique_dates = [
        {"key": date_key, "label": all_dates[date_key]}
        for date_key in sorted(all_dates.keys(), key=lambda value: _parse_date(value) or datetime.max)
    ]

    robot_rows: List[Dict[str, Any]] = []
    robot_summary: Dict[str, Dict[str, int]] = {}

    for robot_id in ROBOT_IDS:
        if robot_id not in robot_events_map:
            continue

        robot_events = robot_events_map[robot_id]
        robot_events.sort(
            key=lambda event: (
                event["date"],
                event.get("report_id") or 0,
                EVENT_TYPE_ORDER.index(event["event_type"]) if event["event_type"] in EVENT_TYPE_ORDER else 999,
            )
        )

        counters = {event_type: 0 for event_type in EVENT_TYPE_ORDER}
        events_by_date: Dict[str, List[Dict[str, Any]]] = {}
        software_state: Dict[str, Dict[str, str]] = {}
        recent_parts: List[str] = []

        for event in robot_events:
            counters[event["event_type"]] += 1
            prefix = EVENT_LABEL_PREFIX.get(event["event_type"], event["event_type"].upper())

            if event["event_type"] == EVENT_TYPE_MANSTRA_SW:
                label = prefix
            elif event["event_type"] == EVENT_TYPE_MANSTRA_HW:
                label = prefix
            else:
                label = f"{prefix}{counters[event['event_type']]}"

            event["label"] = label
            event["robot_url"] = f"/MedicairGeek/infoImpianto/robot/{robot_id}"
            event["detail_modal_title"] = event["titolo"] or event["detail_category"]
            event["detail_modal_text"] = event["detail_text"]
            events_by_date.setdefault(event["date_str"], []).append(event)

        for event in sorted(robot_events, key=lambda item: item["date"], reverse=True):
            for component in event.get("software_components", []):
                if component in software_state:
                    continue
                version = (event.get("software_versions") or [""])[0]
                badge_text = f"{component} {version}".strip()
                software_state[component] = {
                    "text": badge_text,
                    "color": _version_badge_color(version or component),
                }

            for part in event.get("hardware_parts", []):
                if part not in recent_parts:
                    recent_parts.append(part)
                if len(recent_parts) >= 3:
                    break
            if len(recent_parts) >= 3:
                break

        software_badges = list(software_state.values())[:3]
        hardware_badges = [{"text": part} for part in recent_parts[:3]]

        robot_summary[robot_id] = dict(counters)
        robot_rows.append(
            {
                "robot_id": robot_id,
                "robot_url": f"/MedicairGeek/infoImpianto/robot/{robot_id}",
                "events_by_date": events_by_date,
                "software_badges": software_badges,
                "hardware_badges": hardware_badges,
                "summary": robot_summary[robot_id],
            }
        )

    return robot_rows, unique_dates, robot_summary


@disallineamento_qr_bp.get("/MedicairGeek/disallineamentoQr")
@disallineamento_qr_bp.get("/disallineamento-qr")
def disallineamento_qr_page():
    robot_rows, unique_dates, robot_summary = _build_robot_events_table()
    log_activity(f"view | page=disallineamentoQr | ip={request.remote_addr}")

    return render_template(
        "disallineamentoQr.html",
        title="Disallineamento QR",
        now=datetime.now(),
        robot_rows=robot_rows,
        unique_dates=unique_dates,
        robot_summary=robot_summary,
    )


@disallineamento_qr_bp.get("/MedicairGeek/DisallineamentoQR/data")
def api_disallineamento_qr_data():
    robot_rows, unique_dates, robot_summary = _build_robot_events_table()
    out: Dict[str, Any] = {}
    for row in robot_rows:
        out[row["robot_id"]] = {
            "software_badges": row["software_badges"],
            "hardware_badges": row["hardware_badges"],
            "summary": row["summary"],
            "events_by_date": row["events_by_date"],
        }

    return jsonify({
        "ok": True,
        "robots": [row["robot_id"] for row in robot_rows],
        "unique_dates": unique_dates,
        "summary": robot_summary,
        "data": out,
    })
