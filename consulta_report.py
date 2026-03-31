from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import io
import zipfile

from flask import Blueprint, jsonify, render_template, request, abort, send_from_directory, send_file
from openpyxl import load_workbook

from config import REPORT_BASE_DIR


consulta_report_bp = Blueprint("consulta_report", __name__)

BASE_DIR = Path(__file__).resolve().parent
REPORT_DIR = REPORT_BASE_DIR
EXCEL_PATH = REPORT_DIR / "Incidenti_robot.xlsx"

# Estensioni media
IMG_EXT = {"jpg", "jpeg", "png", "gif", "webp", "heic", "bmp", "tiff"}
VID_EXT = {"mp4", "mov", "m4v", "avi", "mkv", "webm"}

SYSTEM_FILENAMES = {"thumbs.db", "desktop.ini", ".ds_store"}

ALLOWED_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".gif", ".webp", ".heic", ".bmp", ".tiff",
    ".mp4", ".mov", ".m4v", ".avi", ".mkv", ".webm",
    ".pdf",
    ".doc", ".docx",
    ".xls", ".xlsx", ".xlsm", ".csv",
    ".ppt", ".pptx",
    ".txt", ".rtf",
}


def _safe_exists_excel() -> bool:
    return EXCEL_PATH.exists() and EXCEL_PATH.is_file()


def _normalize_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _truncate_preview(text: str, max_chars: int = 160) -> str:
    t = _normalize_str(text).replace("\r", "\n")
    t = " ".join(t.split())
    if len(t) <= max_chars:
        return t
    return t[: max_chars - 1].rstrip() + "…"


def _categoria_to_css(categoria: str) -> str:
    c = _normalize_str(categoria).lower()

    if c == "incidente":
        return "incidente"
    if c == "problema software":
        return "problema software"
    if c == "problema hardware":
        return "problema hardware"
    return "altro"


def _find_folder_for_report_id(report_id: int) -> Optional[str]:
    """
    Convenzione cartelle: report/<id>_<dd-mm-YYYY_HH-MM>
    Cerchiamo la cartella che inizia con "{id}_".
    Se ce ne sono più di una, scegliamo la più recente.
    """
    if not REPORT_DIR.exists():
        return None

    prefix = f"{report_id}_"
    candidates: List[Path] = []

    for p in REPORT_DIR.iterdir():
        if p.is_dir() and p.name.startswith(prefix):
            candidates.append(p)

    if not candidates:
        return None

    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return candidates[0].name


def _list_attachments(folder_name: str) -> List[Dict[str, Any]]:
    folder = REPORT_DIR / folder_name
    if not folder.exists() or not folder.is_dir():
        return []

    out: List[Dict[str, Any]] = []

    for p in sorted(folder.iterdir(), key=lambda x: x.name.lower()):
        if not p.is_file():
            continue

        # Evita file di sistema o temporanei
        if p.name.lower() in SYSTEM_FILENAMES or p.name.startswith("~$") or p.name.startswith("."):
            continue

        ext = p.suffix.lower().lstrip(".")
        kind = "file"

        if ext in IMG_EXT:
            kind = "image"
        elif ext in VID_EXT:
            kind = "video"
        elif ext == "pdf":
            kind = "pdf"

        out.append(
            {
                "name": p.name,
                "kind": kind,
                "url": f"/MedicairGeek/ConsultaReport/media/{folder_name}/{p.name}",
            }
        )

    return out


def _get_header_map(ws) -> Dict[str, int]:
    """
    Mappa nome_colonna -> indice colonna (0-based)
    """
    headers: Dict[str, int] = {}

    for idx, cell in enumerate(ws[1]):
        key = _normalize_str(cell.value)
        if key:
            headers[key] = idx

    return headers


def _cell_from_row(row: tuple, header_map: Dict[str, int], key: str) -> Any:
    idx = header_map.get(key)
    if idx is None:
        return None
    if idx >= len(row):
        return None
    return row[idx]


def _read_excel_rows(limit: int = 5000) -> List[Dict[str, Any]]:
    """
    Header attesi:
    id, data, ora, Categoria, Titolo, robot, scaffale, cella, zona,
    luci robot, errore, note, rimosso, risoluzione,
    data_update1, update1, data_update2, update2, codice, sostituito,
    sostituito_qr_scaffale, sostituito_qr_cella, parti_coinvolte
    """
    if not _safe_exists_excel():
        return []

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    header_map = _get_header_map(ws)
    rows: List[Dict[str, Any]] = []

    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue

        raw_id = _cell_from_row(r, header_map, "id")
        if raw_id is None:
            continue

        try:
            rid = int(str(raw_id).strip())
        except Exception:
            continue

        data = _normalize_str(_cell_from_row(r, header_map, "data"))
        ora = _normalize_str(_cell_from_row(r, header_map, "ora"))
        categoria = _normalize_str(_cell_from_row(r, header_map, "Categoria"))
        titolo = _normalize_str(_cell_from_row(r, header_map, "Titolo"))
        robot = _normalize_str(_cell_from_row(r, header_map, "robot"))
        scaffale = _normalize_str(_cell_from_row(r, header_map, "scaffale"))
        cella = _normalize_str(_cell_from_row(r, header_map, "cella"))
        zona = _normalize_str(_cell_from_row(r, header_map, "zona"))
        luci = _normalize_str(_cell_from_row(r, header_map, "luci robot"))
        errore = _normalize_str(_cell_from_row(r, header_map, "errore"))
        note = _normalize_str(_cell_from_row(r, header_map, "note"))
        rimosso = _normalize_str(_cell_from_row(r, header_map, "rimosso"))
        risoluzione = _normalize_str(_cell_from_row(r, header_map, "risoluzione"))

        data_update1 = _normalize_str(_cell_from_row(r, header_map, "data_update1"))
        update1 = _normalize_str(_cell_from_row(r, header_map, "update1"))
        data_update2 = _normalize_str(_cell_from_row(r, header_map, "data_update2"))
        update2 = _normalize_str(_cell_from_row(r, header_map, "update2"))
        codice = _normalize_str(_cell_from_row(r, header_map, "codice"))
        sostituito = _normalize_str(_cell_from_row(r, header_map, "sostituito"))
        sostituito_qr_scaffale = _normalize_str(_cell_from_row(r, header_map, "sostituito_qr_scaffale"))
        sostituito_qr_cella = _normalize_str(_cell_from_row(r, header_map, "sostituito_qr_cella"))
        parti_coinvolte = _normalize_str(_cell_from_row(r, header_map, "parti_coinvolte"))

        has_update = bool(update1 or update2)

        rows.append(
            {
                "id": rid,
                "data": data,
                "ora": ora,
                "dt_label": f"{data} {ora}".strip(),
                "categoria": categoria,
                "categoria_css": _categoria_to_css(categoria),
                "titolo": titolo,
                "robot": robot,
                "scaffale": scaffale,
                "cella": cella,
                "zona": zona,
                "luci": luci,
                "errore": errore,
                "note": note,
                "note_preview": _truncate_preview(note),
                "rimosso": rimosso,
                "risoluzione": risoluzione,
                "data_update1": data_update1,
                "update1": update1,
                "data_update2": data_update2,
                "update2": update2,
                "codice": codice,
                "sostituito": sostituito,
                "sostituito_qr_scaffale": sostituito_qr_scaffale,
                "sostituito_qr_cella": sostituito_qr_cella,
                "parti_coinvolte": parti_coinvolte,
                "has_update": has_update,
            }
        )

        if len(rows) >= limit:
            break

    wb.close()

    def _sort_key(x: Dict[str, Any]) -> Tuple[int, str]:
        try:
            dt = datetime.strptime(f"{x.get('data', '')} {x.get('ora', '')}", "%d/%m/%Y %H:%M")
            return (0, dt.isoformat())
        except Exception:
            return (1, f"{int(x.get('id', 0)):010d}")

    rows.sort(key=_sort_key, reverse=True)
    return rows


def _get_report_by_id(report_id: int) -> Optional[Dict[str, Any]]:
    rows = _read_excel_rows(limit=5000)
    return next((r for r in rows if int(r.get("id", -1)) == int(report_id)), None)


@consulta_report_bp.get("/MedicairGeek/storicoReport")
def storico_report_page():
    return render_template("storicoReport.html", title="Storico Report")


@consulta_report_bp.get("/MedicairGeek/ConsultaReport/list")
def api_list_reports():
    """
    Query params:
    - q: ricerca testo
    - robot: filtro robot
    - categoria: filtro categoria
    - limit: default 300
    """
    q = (request.args.get("q", "") or "").strip().lower()
    robot_filter = (request.args.get("robot", "") or "").strip().lower()
    categoria_filter = (request.args.get("categoria", "") or "").strip().lower()

    try:
        limit = int(request.args.get("limit", "300"))
    except Exception:
        limit = 300

    limit = max(1, min(2000, limit))

    rows = _read_excel_rows(limit=5000)

    if q:
        def match_q(x: Dict[str, Any]) -> bool:
            blob = " ".join(
                [
                    _normalize_str(x.get("titolo")),
                    _normalize_str(x.get("categoria")),
                    _normalize_str(x.get("robot")),
                    _normalize_str(x.get("note")),
                    _normalize_str(x.get("zona")),
                    _normalize_str(x.get("errore")),
                    _normalize_str(x.get("scaffale")),
                    _normalize_str(x.get("cella")),
                    _normalize_str(x.get("update1")),
                    _normalize_str(x.get("update2")),
                ]
            ).lower()
            return q in blob

        rows = [r for r in rows if match_q(r)]

    if robot_filter:
        rows = [r for r in rows if robot_filter in _normalize_str(r.get("robot")).lower()]

    if categoria_filter:
        rows = [r for r in rows if categoria_filter == _normalize_str(r.get("categoria")).lower()]

    out: List[Dict[str, Any]] = []

    for r in rows[:limit]:
        folder = _find_folder_for_report_id(int(r["id"]))
        attachments = _list_attachments(folder) if folder else []

        out.append(
            {
                "id": r["id"],
                "data": r["data"],
                "ora": r["ora"],
                "dt_label": r["dt_label"],
                "titolo": r["titolo"],
                "categoria": r["categoria"],
                "categoria_css": r["categoria_css"],
                "robot": r["robot"],
                "note_preview": r["note_preview"],
                "has_update": bool(r.get("has_update")),
                "has_folder": bool(folder),
                "folder_name": folder or "",
                "attachments_count": len(attachments),
            }
        )

    return jsonify({"ok": True, "reports": out})


@consulta_report_bp.get("/MedicairGeek/ConsultaReport/report/<int:report_id>")
def api_get_report(report_id: int):
    item = _get_report_by_id(report_id)
    if not item:
        return jsonify({"ok": False, "error": "Report non trovato."}), 404

    folder = _find_folder_for_report_id(report_id)
    attachments = _list_attachments(folder) if folder else []

    return jsonify(
        {
            "ok": True,
            "report": {
                "id": item["id"],
                "data": item["data"],
                "ora": item["ora"],
                "dt_label": item["dt_label"],
                "titolo": item["titolo"],
                "categoria": item["categoria"],
                "categoria_css": item["categoria_css"],
                "robot": item["robot"],
                "scaffale": item["scaffale"],
                "cella": item["cella"],
                "zona": item["zona"],
                "luci": item["luci"],
                "errore": item["errore"],
                "note": item["note"],
                "rimosso": item["rimosso"],
                "risoluzione": item["risoluzione"],
                "data_update1": item["data_update1"],
                "update1": item["update1"],
                "data_update2": item["data_update2"],
                "update2": item["update2"],
                "codice": item.get("codice", ""),
                "sostituito": item.get("sostituito", ""),
                "sostituito_qr_scaffale": item.get("sostituito_qr_scaffale", ""),
                "sostituito_qr_cella": item.get("sostituito_qr_cella", ""),
                "parti_coinvolte": item.get("parti_coinvolte", ""),
                "has_update": bool(item.get("has_update")),
                "folder_name": folder or "",
                "attachments": attachments,
            },
        }
    )


@consulta_report_bp.get("/MedicairGeek/ConsultaReport/media/<path:folder_name>/<path:filename>")
def serve_report_media(folder_name: str, filename: str):
    """
    Serve in modo sicuro i file dentro report/<folder_name>/<filename>.
    Blocca path traversal.
    """
    base = REPORT_DIR.resolve()
    folder = (REPORT_DIR / folder_name).resolve()

    if base not in folder.parents and folder != base:
        abort(404)

    file_path = (folder / filename).resolve()
    if folder not in file_path.parents:
        abort(404)

    if not file_path.exists() or not file_path.is_file():
        abort(404)

    return send_from_directory(directory=str(folder), path=filename, as_attachment=False)


@consulta_report_bp.get("/MedicairGeek/storicoReport/<int:report_id>")
def storico_report_detail_page(report_id: int):
    return render_template(
        "reportDetail.html",
        title=f"Dettaglio Report #{report_id}",
        report_id=report_id,
    )


@consulta_report_bp.get("/MedicairGeek/ConsultaReport/download/<int:report_id>")
def download_report_files(report_id: int):
    report_root = REPORT_DIR

    if not report_root.exists() or not report_root.is_dir():
        abort(500, description="Directory report non disponibile.")

    matches = [
        p for p in report_root.iterdir()
        if p.is_dir() and p.name.startswith(f"{report_id}_")
    ]
    if not matches:
        abort(404, description="Cartella report non trovata.")

    matches.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    folder = matches[0]

    files = [
        p for p in folder.iterdir()
        if (
            p.is_file()
            and p.name.lower() not in SYSTEM_FILENAMES
            and not p.name.startswith("~$")
            and not p.name.startswith(".")
            and p.suffix.lower() in ALLOWED_EXTENSIONS
        )
    ]

    files.sort(key=lambda p: p.name.lower())

    if not files:
        abort(404, description="Nessun file valido presente nel report.")

    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path in files:
            zf.write(file_path, arcname=file_path.name)

    memory_file.seek(0)

    return send_file(
        memory_file,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"report_{report_id}.zip",
    )