"""Blueprint per la pagina Disallineamento QR.

Mostra una tabella con i robot nella prima colonna e le date degli eventi nelle colonne successive.
Ogni evento ha un simbolo e un colore diverso:
- QR: Disallineamento QR (rosso)
- FW: Update Firmware (blu)
- MCU: Cambio MCU (giallo)
- BAT: Cambio Batteria (verde)
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from flask import Blueprint, render_template, request, jsonify
from openpyxl import load_workbook

from config import REPORT_BASE_DIR
from log_utils import log_activity


disallineamento_qr_bp = Blueprint("disallineamento_qr", __name__)

REPORT_DIR = REPORT_BASE_DIR
EXCEL_PATH = REPORT_DIR / "Incidenti_robot.xlsx"

# Registro dei robot
ROBOT_IDS = ["16278", "16279", "16292", "16294", "16302", "16306", 
             "16313", "16314", "16325", "16337", "16339", "16340", "16348", "16349", "16350"]

# Const per i tipi di evento
EVENT_TYPE_QR = "qr"
EVENT_TYPE_FW = "fw"
EVENT_TYPE_MCU = "mcu"
EVENT_TYPE_BAT = "bat"


def _normalize_str(v: Any) -> str:
    """Normalizza un valore a stringa."""
    if v is None:
        return ""
    return str(v).strip()


def _parse_date(date_str: str) -> datetime | None:
    """Tenta di parsare una data nel formato GG/MM/YYYY."""
    try:
        d = _normalize_str(date_str)
        if not d:
            return None
        return datetime.strptime(d, "%d/%m/%Y")
    except Exception:
        return None


def _categorize_event(categoria: str, titolo: str, parti_coinvolte: str) -> str | None:
    """Categorizza un evento in uno dei 4 tipi.
    
    Ritorna:
    - "qr" per disallineamento QR
    - "fw" per update firmware
    - "mcu" per cambio MCU
    - "bat" per cambio batteria
    - None se non corrisponde a nessuna categoria
    """
    categoria_lower = _normalize_str(categoria).lower()
    titolo_lower = _normalize_str(titolo).lower()
    parti_lower = _normalize_str(parti_coinvolte).lower()
    
    # Disallineamento QR
    if "dissalineato" in categoria_lower and "qr" in categoria_lower:
        return EVENT_TYPE_QR
    
    # Interventi di manutenzione
    if "intervento manutenzione" in categoria_lower:
        # Cambio Batteria
        if "batteria" in titolo_lower or "batteria" in parti_lower:
            return EVENT_TYPE_BAT
        # Firmware Update
        if "firm" in titolo_lower or "firm" in parti_lower or "firmare" in titolo_lower:
            return EVENT_TYPE_FW
        # Cambio MCU
        if "scheda madre" in titolo_lower or "scheda madre" in parti_lower or "mcu" in titolo_lower or "mcu" in parti_lower:
            return EVENT_TYPE_MCU
    
    return None


def _read_events_from_excel() -> List[Dict[str, Any]]:
    """Legge gli eventi dal file Excel e ritorna una lista filtrata.
    
    Ritorna una lista di eventi con:
    - data: data come datetime
    - date_str: data come stringa GG/MM/YYYY
    - robots: lista di ID robot (possono essere multiple in una sola riga)
    - event_type: tipo di evento (qr, fw, mcu, bat)
    - titolo: titolo dell'evento
    """
    events: List[Dict[str, Any]] = []
    
    if not EXCEL_PATH.exists() or not EXCEL_PATH.is_file():
        return events
    
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        
        # Mappa colonne
        header_map = {}
        for idx, cell in enumerate(ws[1]):
            header_map[_normalize_str(cell.value)] = idx
        
        # Leggi tutte le righe
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            
            # Estrai i dati
            data_str = _normalize_str(row[header_map.get("data", 1)])
            categoria = _normalize_str(row[header_map.get("Categoria", 3)])
            titolo = _normalize_str(row[header_map.get("Titolo", 4)])
            robot_str = _normalize_str(row[header_map.get("robot", 5)])
            parti_coinvolte = _normalize_str(row[header_map.get("parti_coinvolte", 21)])
            
            # Parsa la data
            date_obj = _parse_date(data_str)
            if not date_obj:
                continue
            
            # Categorizza l'evento
            event_type = _categorize_event(categoria, titolo, parti_coinvolte)
            if not event_type:
                continue
            
            # Parsa i robot (possono essere separati da virgola o spazio)
            robots = []
            for robot_raw in robot_str.split(","):
                robot_clean = _normalize_str(robot_raw)
                # Cerca i robot con ID simile (potrebbe contenere altri dati)
                for robot_id in ROBOT_IDS:
                    if robot_id in robot_clean:
                        robots.append(robot_id)
            
            # Se non ha robot validi, salta
            if not robots:
                continue
            
            # Aggiungi all'elenco
            events.append({
                "date": date_obj,
                "date_str": data_str,
                "robots": robots,
                "event_type": event_type,
                "titolo": titolo,
            })
        
        wb.close()
    except Exception as exc:
        print(f"[DisallineamentoQR] Errore lettura Excel: {exc}")
    
    return events


def _build_robot_events_table() -> tuple[List[str], Dict[str, List[Dict[str, Any]]], List[str], Dict[str, Dict[str, int]]]:
    """Costruisce la tabella degli eventi per i robot.
    
    Ritorna:
    - robots_list: lista ordinata degli ID robot che hanno eventi
    - robot_data: dict {robot_id: [{date_str, date, event_type, titolo, label}, ...]}
    - unique_dates_sorted: lista di tutte le date uniche ordinate (più recenti prima)
    - robot_summary: dict {robot_id: {fw: count, mcu: count, bat: count}}
    """
    events = _read_events_from_excel()
    
    # Organizza per robot
    robot_data: Dict[str, List[Dict[str, Any]]] = {}
    all_dates = set()
    
    for event in events:
        for robot_id in event["robots"]:
            if robot_id not in robot_data:
                robot_data[robot_id] = []
            
            robot_data[robot_id].append({
                "date_str": event["date_str"],
                "date": event["date"],
                "event_type": event["event_type"],
                "titolo": event["titolo"],
            })
            
            all_dates.add(event["date_str"])
    
    # Ordina i robot in base all'ordine definito
    robots_list = [rid for rid in ROBOT_IDS if rid in robot_data]
    
    # Ordina le date (più recenti prima)
    unique_dates_sorted = sorted(list(all_dates), 
                                 key=lambda d: _parse_date(d), 
                                 reverse=True)
    
    robot_summary: Dict[str, Dict[str, int]] = {}
    
    # Ordina gli eventi per ogni robot per data e assegna numeri progressivi per tipo
    for robot_id, robot_events in robot_data.items():
        robot_events.sort(key=lambda e: e["date"])
        counters = {EVENT_TYPE_QR: 0, EVENT_TYPE_FW: 0, EVENT_TYPE_MCU: 0, EVENT_TYPE_BAT: 0}
        
        for event in robot_events:
            counters[event["event_type"]] += 1
            label = event["event_type"].upper()
            if event["event_type"] == EVENT_TYPE_QR:
                label = f"QR{counters[event['event_type']]}"
            elif event["event_type"] == EVENT_TYPE_FW:
                label = f"FW{counters[event['event_type']]}"
            elif event["event_type"] == EVENT_TYPE_MCU:
                label = f"MCU{counters[event['event_type']]}"
            elif event["event_type"] == EVENT_TYPE_BAT:
                label = f"BAT{counters[event['event_type']]}"
            event["label"] = label
        
        # Salva il riepilogo dei tipi per la riga robot
        robot_summary[robot_id] = {
            EVENT_TYPE_FW: counters[EVENT_TYPE_FW],
            EVENT_TYPE_MCU: counters[EVENT_TYPE_MCU],
            EVENT_TYPE_BAT: counters[EVENT_TYPE_BAT],
        }
        
        # Ordina gli eventi per visualizzazione in tabella (più recenti prima)
        robot_events.sort(key=lambda e: e["date"], reverse=True)
    
    return robots_list, robot_data, unique_dates_sorted, robot_summary


@disallineamento_qr_bp.get("/MedicairGeek/disallineamentoQr")
@disallineamento_qr_bp.get("/disallineamento-qr")
def disallineamento_qr_page():
    """Pagina principale disallineamento QR."""
    robots_list, robot_data, unique_dates, robot_summary = _build_robot_events_table()
    log_activity(f"view | page=disallineamentoQr | ip={request.remote_addr}")
    
    return render_template(
        "disallineamentoQr.html",
        title="Disallineamento QR",
        now=datetime.now(),
        robots_list=robots_list,
        robot_data=robot_data,
        unique_dates=unique_dates,
        robot_summary=robot_summary,
    )


@disallineamento_qr_bp.get("/MedicairGeek/DisallineamentoQR/data")
def api_disallineamento_qr_data():
    """API per ottenere i dati in JSON."""
    robots_list, robot_data, unique_dates, robot_summary = _build_robot_events_table()
    
    # Formatta per JSON
    out = {}
    for robot_id in robots_list:
        out[robot_id] = []
        if robot_id in robot_data:
            for event in robot_data[robot_id]:
                out[robot_id].append({
                    "date": event["date_str"],
                    "event_type": event["event_type"],
                    "label": event.get("label", ""),
                    "titolo": event["titolo"],
                })
    
    return jsonify({
        "ok": True,
        "robots": robots_list,
        "unique_dates": unique_dates,
        "summary": robot_summary,
        "data": out
    })
