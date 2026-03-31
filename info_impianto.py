from __future__ import annotations

import json
import re
import threading
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from flask import Blueprint, abort, render_template, request
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from config import REPORT_INFO_IMPIANTO_DIR
from consulta_report import _read_excel_rows


info_impianto_bp = Blueprint("info_impianto", __name__)

INFO_IMPIANTO_DIR = REPORT_INFO_IMPIANTO_DIR
INDEX_CACHE_PATH = INFO_IMPIANTO_DIR / "00_indice_info_impianto.xlsx"
CACHE_MANIFEST_PATH = INFO_IMPIANTO_DIR / "_cache_manifest.json"
CACHE_VERSION = 1
CACHE_LOCK = threading.RLock()

ROBOT_REGISTRY: List[Dict[str, str]] = [
    {"id": "16278", "full_id": "1216278", "ip": "10.1.70.57/24", "label": "Robot 16278"},
    {"id": "16279", "full_id": "1216279", "ip": "10.1.70.55/24", "label": "Robot 16279"},
    {"id": "16292", "full_id": "1216292", "ip": "10.1.70.60/24", "label": "Robot 16292"},
    {"id": "16294", "full_id": "1216294", "ip": "10.1.70.62/24", "label": "Robot 16294"},
    {"id": "16302", "full_id": "1216302", "ip": "10.1.70.53/24", "label": "Robot 16302"},
    {"id": "16306", "full_id": "1216306", "ip": "10.1.70.56/24", "label": "Robot 16306"},
    {"id": "16313", "full_id": "1216313", "ip": "10.1.70.54/24", "label": "Robot 16313"},
    {"id": "16314", "full_id": "1216314", "ip": "10.1.70.51/24", "label": "Robot 16314"},
    {"id": "16325", "full_id": "1216325", "ip": "10.1.70.61/24", "label": "Robot 16325"},
    {"id": "16337", "full_id": "1216337", "ip": "10.1.70.52/24", "label": "Robot 16337"},
    {"id": "16339", "full_id": "1216339", "ip": "10.1.70.63/24", "label": "Robot 16339"},
    {"id": "16340", "full_id": "1216340", "ip": "10.1.70.64/24", "label": "Robot 16340"},
    {"id": "16348", "full_id": "1216348", "ip": "10.1.70.58/24", "label": "Robot 16348"},
    {"id": "16349", "full_id": "1216349", "ip": "10.1.70.50/24", "label": "Robot 16349"},
    {"id": "16350", "full_id": "1216350", "ip": "10.1.70.59/24", "label": "Robot 16350"},
]

HARDWARE_ITEMS: List[Dict[str, str]] = [
    {"label": "Pavimento", "slug": "pavimento", "meta": "QR sostituiti e report"},
    {"label": "Working Station 1", "slug": "workingstation1", "meta": "PC 10.1.70.80 • Label 10.1.70.90"},
    {"label": "Working Station 2", "slug": "workingstation2", "meta": "PC 10.1.70.81 • Label 10.1.70.91"},
    {"label": "Working Station 3", "slug": "workingstation3", "meta": "PC 10.1.70.82 • Label 10.1.70.92"},
    {"label": "Charging Station 1", "slug": "chargingstation1", "meta": "IP 10.1.70.100"},
    {"label": "Charging Station 2", "slug": "chargingstation2", "meta": "IP 10.1.70.101"},
    {"label": "Scaffale", "slug": "scaffale", "meta": "QR sostituiti e report"},
]

SOFTWARE_ITEMS: List[Dict[str, str]] = [
    {"label": "Rete Wifi/LAN", "slug": "rete-wifi-lan", "meta": "Elenco IP impianto"},
    {"label": "RMS", "slug": "rms", "meta": "Storico report RMS"},
    {"label": "WMS", "slug": "wms", "meta": "Storico report WMS"},
    {"label": "Altro", "slug": "altro", "meta": "Segnalazioni generiche"},
]

NETWORK_IP_ROWS: List[Dict[str, str]] = [
    {"name": "Robot 16278 (1216278)", "ip": "10.1.70.57"},
    {"name": "Robot 16279 (1216279)", "ip": "10.1.70.55"},
    {"name": "Robot 16292 (1216292)", "ip": "10.1.70.60"},
    {"name": "Robot 16294 (1216294)", "ip": "10.1.70.62"},
    {"name": "Robot 16302 (1216302)", "ip": "10.1.70.53"},
    {"name": "Robot 16306 (1216306)", "ip": "10.1.70.56"},
    {"name": "Robot 16313 (1216313)", "ip": "10.1.70.54"},
    {"name": "Robot 16314 (1216314)", "ip": "10.1.70.51"},
    {"name": "Robot 16325 (1216325)", "ip": "10.1.70.61"},
    {"name": "Robot 16337 (1216337)", "ip": "10.1.70.52"},
    {"name": "Robot 16339 (1216339)", "ip": "10.1.70.63"},
    {"name": "Robot 16340 (1216340)", "ip": "10.1.70.64"},
    {"name": "Robot 16348 (1216348)", "ip": "10.1.70.58"},
    {"name": "Robot 16349 (1216349)", "ip": "10.1.70.50"},
    {"name": "Robot 16350 (1216350)", "ip": "10.1.70.59"},
    {"name": "Working Station 1 - PC", "ip": "10.1.70.80"},
    {"name": "Working Station 1 - Label", "ip": "10.1.70.90"},
    {"name": "Working Station 2 - PC", "ip": "10.1.70.81"},
    {"name": "Working Station 2 - Label", "ip": "10.1.70.91"},
    {"name": "Working Station 3 - PC", "ip": "10.1.70.82"},
    {"name": "Working Station 3 - Label", "ip": "10.1.70.92"},
    {"name": "Charging Station 1", "ip": "10.1.70.100"},
    {"name": "Charging Station 2", "ip": "10.1.70.101"},
]

COMPONENT_REGISTRY: List[Dict[str, Any]] = [
    {
        "slug": "pavimento",
        "label": "Pavimento",
        "section": "hardware",
        "description": "Area impianto con storico report e QR sostituiti.",
        "aliases": ["pavimento"],
        "show_qr_table": True,
        "info_cards": [{"label": "Tipologia", "value": "Area di passaggio / impianto"}],
    },
    {
        "slug": "workingstation1",
        "label": "Working Station 1",
        "section": "hardware",
        "description": "Postazione di lavoro con PC e stampante label.",
        "aliases": ["workingstation1", "working station 1", "station 1"],
        "info_cards": [
            {"label": "IP PC", "value": "10.1.70.80"},
            {"label": "IP Label", "value": "10.1.70.90"},
        ],
    },
    {
        "slug": "workingstation2",
        "label": "Working Station 2",
        "section": "hardware",
        "description": "Postazione di lavoro con PC e stampante label.",
        "aliases": ["workingstation2", "working station 2", "station 2"],
        "info_cards": [
            {"label": "IP PC", "value": "10.1.70.81"},
            {"label": "IP Label", "value": "10.1.70.91"},
        ],
    },
    {
        "slug": "workingstation3",
        "label": "Working Station 3",
        "section": "hardware",
        "description": "Postazione di lavoro con PC e stampante label.",
        "aliases": ["workingstation3", "working station 3", "station 3"],
        "info_cards": [
            {"label": "IP PC", "value": "10.1.70.82"},
            {"label": "IP Label", "value": "10.1.70.92"},
        ],
    },
    {
        "slug": "chargingstation1",
        "label": "Charging Station 1",
        "section": "hardware",
        "description": "Stazione di ricarica robot.",
        "aliases": ["chargingstation1", "charging station 1"],
        "info_cards": [{"label": "IP", "value": "10.1.70.100"}],
    },
    {
        "slug": "chargingstation2",
        "label": "Charging Station 2",
        "section": "hardware",
        "description": "Stazione di ricarica robot.",
        "aliases": ["chargingstation2", "charging station 2"],
        "info_cards": [{"label": "IP", "value": "10.1.70.101"}],
    },
    {
        "slug": "scaffale",
        "label": "Scaffale",
        "section": "hardware",
        "description": "Dettaglio scaffali con storico report e QR sostituiti.",
        "aliases": ["scaffale", "scaffali"],
        "show_qr_table": True,
        "info_cards": [{"label": "Tipologia", "value": "Area scaffalatura"}],
    },
    {
        "slug": "rete-wifi-lan",
        "label": "Rete Wifi / LAN",
        "section": "software",
        "description": "Riepilogo IP noti dell'impianto e storico report rete.",
        "aliases": ["rete wifi/lan", "rete wifi", "wifi", "lan"],
        "info_cards": [{"label": "Copertura", "value": "IP impianto disponibili sotto"}],
        "network_rows": NETWORK_IP_ROWS,
    },
    {
        "slug": "rms",
        "label": "RMS",
        "section": "software",
        "description": "Storico segnalazioni e report RMS.",
        "aliases": ["rms"],
        "info_cards": [{"label": "Sistema", "value": "RMS"}],
    },
    {
        "slug": "wms",
        "label": "WMS",
        "section": "software",
        "description": "Storico segnalazioni e report WMS.",
        "aliases": ["wms"],
        "info_cards": [{"label": "Sistema", "value": "WMS"}],
    },
    {
        "slug": "altro",
        "label": "Altro",
        "section": "software",
        "description": "Contenitore per segnalazioni generiche non classificate.",
        "aliases": ["altro"],
        "info_cards": [{"label": "Categoria", "value": "Segnalazioni varie"}],
    },
]


def _normalize_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _truncate(text: str, max_chars: int = 120) -> str:
    clean = " ".join(_normalize_str(text).replace("\r", " ").replace("\n", " ").split())
    if len(clean) <= max_chars:
        return clean
    return clean[: max_chars - 1].rstrip() + "…"


def _pick_first_non_empty(*values: Any) -> str:
    for value in values:
        text = _normalize_str(value)
        if text:
            return text
    return ""


def _clean_ip(value: Any) -> str:
    return _normalize_str(value).split("/", 1)[0]


def _is_yes_value(value: Any) -> bool:
    return _normalize_str(value).lower() in {"si", "sì", "yes", "y", "true", "1"}


def _get_robot(robot_id: str) -> Optional[Dict[str, str]]:
    key = _normalize_str(robot_id)
    for robot in ROBOT_REGISTRY:
        if robot["id"] == key or robot["full_id"] == key:
            return robot
    return None


def _get_component(component_slug: str) -> Optional[Dict[str, Any]]:
    key = _normalize_str(component_slug).lower()
    for component in COMPONENT_REGISTRY:
        if _normalize_str(component.get("slug")).lower() == key:
            return component
    return None


def _robot_aliases(robot: Dict[str, str]) -> set[str]:
    short_id = robot.get("id", "")
    full_id = robot.get("full_id", "") or (f"12{short_id}" if short_id else "")
    return {
        short_id,
        full_id,
        f"Robot {short_id}",
        f"Robot {full_id}",
    }


def _robot_matches(report_robot_value: Any, robot: Dict[str, str]) -> bool:
    value = _normalize_str(report_robot_value).lower()
    if not value:
        return False

    if "tutti" in value:
        return True

    aliases = {alias.lower() for alias in _robot_aliases(robot) if alias}
    return any(alias in value for alias in aliases)


def _is_global_robot_report(row: Dict[str, Any]) -> bool:
    return "tutti" in _normalize_str(row.get("robot")).lower()


def _build_parti_coinvolte(row: Dict[str, Any]) -> str:
    explicit_parts = _normalize_str(row.get("parti_coinvolte"))
    if explicit_parts:
        return _truncate(explicit_parts, 120)

    blocks: List[str] = []

    errore = _normalize_str(row.get("errore"))
    risoluzione = _normalize_str(row.get("risoluzione"))
    zona = _normalize_str(row.get("zona"))
    cella = _normalize_str(row.get("cella"))
    scaffale = _normalize_str(row.get("scaffale"))

    if errore:
        blocks.append(_truncate(errore, 90))
    elif risoluzione:
        blocks.append(_truncate(risoluzione, 90))

    area_parts = [x for x in [zona, cella, scaffale] if x and x.lower() != "senza scaffale"]
    if area_parts:
        blocks.append("Area: " + " / ".join(area_parts))

    if not blocks:
        return "-"
    return " | ".join(blocks[:2])


def _extract_replaced_part(row: Dict[str, Any]) -> str:
    explicit_parts = _normalize_str(row.get("parti_coinvolte"))
    if explicit_parts:
        return _truncate(explicit_parts, 120)

    candidates = [
        row.get("risoluzione"),
        row.get("note"),
        row.get("update1"),
        row.get("update2"),
        row.get("errore"),
    ]
    keywords = ("sostit", "cambiat", "ricambio", "rimpiazz", "sostitu")

    for value in candidates:
        text = _normalize_str(value)
        if text and any(keyword in text.lower() for keyword in keywords):
            return _truncate(text, 120)

    categoria = _normalize_str(row.get("categoria")).lower()
    risoluzione = _normalize_str(row.get("risoluzione"))
    if categoria == "intervento manutenzione" and risoluzione:
        return _truncate(risoluzione, 120)

    return ""


def _get_related_rows_for_robot(robot: Dict[str, str], rows: Optional[List[Dict[str, Any]]] = None) -> List[Dict[str, Any]]:
    base_rows = rows if rows is not None else _read_excel_rows(limit=5000)
    return [row for row in base_rows if _robot_matches(row.get("robot"), robot)]


def _decorate_robot(robot: Dict[str, str], related_rows: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
    related_rows = related_rows if related_rows is not None else _get_related_rows_for_robot(robot)
    status_row = next((row for row in related_rows if not _is_global_robot_report(row)), {})
    is_removed = _is_yes_value(status_row.get("rimosso")) if status_row else False

    decorated: Dict[str, Any] = dict(robot)
    decorated["clean_ip"] = _clean_ip(robot.get("ip", ""))
    decorated["display_id"] = f"{robot.get('id', '')} ({robot.get('full_id', '')})" if robot.get("full_id") else robot.get("id", "")
    decorated["is_removed"] = is_removed
    decorated["status_label"] = "Robot rimosso = SI" if is_removed else "Robot rimosso = NO"
    decorated["latest_report_label"] = _pick_first_non_empty(status_row.get("dt_label"), status_row.get("data"), "")
    return decorated


def _build_robot_list() -> List[Dict[str, Any]]:
    rows = _read_excel_rows(limit=5000)
    return [_decorate_robot(robot, _get_related_rows_for_robot(robot, rows)) for robot in ROBOT_REGISTRY]


def _component_aliases(component: Dict[str, Any]) -> set[str]:
    aliases: set[str] = set()

    for raw in [component.get("label"), component.get("slug"), *(component.get("aliases") or [])]:
        text = _normalize_str(raw).lower()
        if not text:
            continue
        aliases.add(text)
        aliases.add(text.replace(" ", ""))

    return aliases


def _component_matches(report_robot_value: Any, component: Dict[str, Any]) -> bool:
    value = _normalize_str(report_robot_value).lower()
    if not value:
        return False

    compact_value = value.replace(" ", "")
    return any(alias and (alias in value or alias in compact_value) for alias in _component_aliases(component))


def _get_related_rows_for_component(component: Dict[str, Any], rows: Optional[List[Dict[str, Any]]] = None) -> List[Dict[str, Any]]:
    base_rows = rows if rows is not None else _read_excel_rows(limit=5000)
    return [row for row in base_rows if _component_matches(row.get("robot"), component)]


def _build_component_tables(component: Dict[str, Any], related_rows: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
    related_rows = related_rows if related_rows is not None else _get_related_rows_for_component(component)

    reports: List[Dict[str, Any]] = []
    qr_changes: List[Dict[str, Any]] = []

    for row in related_rows:
        report_id = int(row.get("id", 0))
        link = f"/MedicairGeek/storicoReport/{report_id}" if report_id else "#"
        data_label = _pick_first_non_empty(row.get("dt_label"), row.get("data"), "-")
        descrizione = _pick_first_non_empty(row.get("note"), row.get("risoluzione"), row.get("errore"), "-")

        reports.append(
            {
                "data": data_label,
                "titolo": _pick_first_non_empty(row.get("titolo"), "-"),
                "tipo": _pick_first_non_empty(row.get("categoria"), "-"),
                "descrizione": _truncate(descrizione, 140) if descrizione != "-" else "-",
                "rimosso": _pick_first_non_empty(row.get("rimosso"), "-"),
                "link": link,
            }
        )

        if component.get("show_qr_table"):
            component_slug = _normalize_str(component.get("slug"))
            legacy_codice = _normalize_str(row.get("codice"))
            legacy_sostituito = _is_yes_value(row.get("sostituito"))

            if component_slug == "scaffale":
                riferimento = _normalize_str(row.get("scaffale"))
                qr_values = _normalize_str(row.get("sostituito_qr_scaffale"))
            else:
                riferimento = _pick_first_non_empty(row.get("cella"), row.get("scaffale"))
                qr_values = _normalize_str(row.get("sostituito_qr_cella"))

            if not qr_values and legacy_codice and legacy_sostituito:
                qr_values = legacy_codice

            if riferimento.lower() == "senza scaffale":
                riferimento = ""

            if qr_values:
                qr_changes.append(
                    {
                        "data": data_label,
                        "riferimento": riferimento or "-",
                        "sostituito_qr": qr_values,
                        "descrizione": _truncate(descrizione, 120) if descrizione != "-" else "-",
                        "link": link,
                    }
                )

    return {
        "reports": reports[:100],
        "qr_changes": qr_changes[:100],
        "related_count": len(related_rows),
    }


def _build_robot_tables(robot: Dict[str, str], related_rows: Optional[List[Dict[str, Any]]] = None) -> Dict[str, List[Dict[str, Any]]]:
    related_rows = related_rows if related_rows is not None else _get_related_rows_for_robot(robot)

    replaced_parts: List[Dict[str, Any]] = []
    manutenzioni: List[Dict[str, Any]] = []
    incidenti: List[Dict[str, Any]] = []

    for row in related_rows:
        report_id = int(row.get("id", 0))
        link = f"/MedicairGeek/storicoReport/{report_id}" if report_id else "#"
        data_label = _pick_first_non_empty(row.get("dt_label"), row.get("data"), "-")
        descrizione = _pick_first_non_empty(row.get("note"), row.get("risoluzione"), row.get("errore"), "-")

        incidenti.append(
            {
                "data": data_label,
                "titolo": _pick_first_non_empty(row.get("titolo"), "-"),
                "tipo": _pick_first_non_empty(row.get("categoria"), "-"),
                "descrizione": _truncate(descrizione, 140) if descrizione != "-" else "-",
                "rimosso": _pick_first_non_empty(row.get("rimosso"), "-"),
                "link": link,
            }
        )

        if _normalize_str(row.get("categoria")).lower() == "intervento manutenzione":
            manutenzioni.append(
                {
                    "data": data_label,
                    "titolo": _pick_first_non_empty(row.get("titolo"), "-"),
                    "parti_coinvolte": _build_parti_coinvolte(row),
                    "descrizione": _truncate(descrizione, 140) if descrizione != "-" else "-",
                    "link": link,
                }
            )

        part_text = _extract_replaced_part(row)
        if part_text:
            replaced_parts.append(
                {
                    "data": data_label,
                    "pezzo": part_text,
                    "link": link,
                }
            )

    return {
        "replaced_parts": replaced_parts[:50],
        "manutenzioni": manutenzioni[:50],
        "incidenti": incidenti[:100],
        "related_count": len(related_rows),
    }


def _safe_slug(value: Any) -> str:
    text = re.sub(r"[^a-zA-Z0-9]+", "_", _normalize_str(value)).strip("_").lower()
    return text or "item"


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(str(value).strip()))
    except Exception:
        return default


def _robot_cache_path(robot: Dict[str, str]) -> Path:
    return INFO_IMPIANTO_DIR / f"robot_{robot.get('id', '')}_{robot.get('full_id', '')}.xlsx"


def _component_cache_path(component: Dict[str, Any]) -> Path:
    prefix = _safe_slug(component.get("section") or "item")
    slug = _safe_slug(component.get("slug"))
    return INFO_IMPIANTO_DIR / f"{prefix}_{slug}.xlsx"


def _ensure_cache_dir() -> None:
    INFO_IMPIANTO_DIR.mkdir(parents=True, exist_ok=True)


def _sheet_to_dict_rows(ws) -> List[Dict[str, str]]:
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return []

    headers = [_normalize_str(value) for value in values[0]]
    rows: List[Dict[str, str]] = []

    for raw_row in values[1:]:
        row: Dict[str, str] = {}
        has_value = False
        for idx, header in enumerate(headers):
            if not header:
                continue
            cell_value = raw_row[idx] if idx < len(raw_row) else ""
            text = _normalize_str(cell_value)
            row[header] = text
            if text:
                has_value = True
        if has_value:
            rows.append(row)

    return rows


def _sheet_to_key_value_map(ws) -> Dict[str, str]:
    info: Dict[str, str] = {}
    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        if not raw_row:
            continue
        key = _normalize_str(raw_row[0] if len(raw_row) > 0 else "")
        value = _normalize_str(raw_row[1] if len(raw_row) > 1 else "")
        if key:
            info[key] = value
    return info


def _get_manual_rows(path: Path) -> List[List[Any]]:
    if not path.exists():
        return []

    try:
        wb = load_workbook(path, data_only=False)
        if "Note manuali" not in wb.sheetnames:
            wb.close()
            return []

        ws = wb["Note manuali"]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    except Exception:
        return []


def _autosize_worksheet(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_index = column_cells[0].column
        for cell in column_cells:
            try:
                max_length = max(max_length, len(str(cell.value or "")))
            except Exception:
                continue
        ws.column_dimensions[get_column_letter(column_index)].width = max(12, min(max_length + 2, 60))


def _append_rows_sheet(wb: Workbook, title: str, headers: List[str], rows: List[List[Any]]) -> None:
    ws = wb.create_sheet(title=title[:31])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _autosize_worksheet(ws)


def _append_manual_sheet(wb: Workbook, manual_rows: List[List[Any]]) -> None:
    ws = wb.create_sheet(title="Note manuali")
    if manual_rows:
        for row in manual_rows:
            ws.append(list(row))
    else:
        ws.append(["Campo", "Valore", "Note"])
        ws.append(["", "", "Questa scheda viene preservata nei refresh automatici."])
    _autosize_worksheet(ws)


def _write_robot_cache(robot: Dict[str, str], related_rows: List[Dict[str, Any]], exported_at: str) -> None:
    robot_view = _decorate_robot(robot, related_rows)
    tables = _build_robot_tables(robot, related_rows)
    cache_path = _robot_cache_path(robot)
    manual_rows = _get_manual_rows(cache_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Info"
    ws.append(["Campo", "Valore"])
    info_rows = [
        ["Robot", robot_view.get("label", "")],
        ["ID breve", robot_view.get("id", "")],
        ["ID completo", robot_view.get("full_id", "")],
        ["ID display", robot_view.get("display_id", "")],
        ["IP", robot_view.get("clean_ip", "")],
        ["Stato rimosso", "SI" if robot_view.get("is_removed") else "NO"],
        ["Status label", robot_view.get("status_label", "")],
        ["Ultimo report", robot_view.get("latest_report_label") or "-"],
        ["Report collegati", tables.get("related_count", 0)],
        ["Aggiornato il", exported_at],
        ["File Excel", cache_path.name],
        ["Link app", f"/MedicairGeek/infoImpianto/robot/{robot.get('id', '')}"],
    ]
    for row in info_rows:
        ws.append(row)
    _autosize_worksheet(ws)

    _append_rows_sheet(
        wb,
        "Pezzi sostituiti",
        ["Data", "Pezzo", "Link app"],
        [[item.get("data", "-"), item.get("pezzo", "-"), item.get("link", "#")] for item in tables.get("replaced_parts", [])],
    )
    _append_rows_sheet(
        wb,
        "Manutenzioni",
        ["Data", "Titolo", "Parti coinvolte", "Descrizione", "Link app"],
        [
            [
                item.get("data", "-"),
                item.get("titolo", "-"),
                item.get("parti_coinvolte", "-"),
                item.get("descrizione", "-"),
                item.get("link", "#"),
            ]
            for item in tables.get("manutenzioni", [])
        ],
    )
    _append_rows_sheet(
        wb,
        "Incidenti",
        ["Data", "Titolo", "Tipo", "Descrizione", "Rimosso", "Link app"],
        [
            [
                item.get("data", "-"),
                item.get("titolo", "-"),
                item.get("tipo", "-"),
                item.get("descrizione", "-"),
                item.get("rimosso", "-"),
                item.get("link", "#"),
            ]
            for item in tables.get("incidenti", [])
        ],
    )
    _append_manual_sheet(wb, manual_rows)
    wb.save(cache_path)
    wb.close()


def _write_component_cache(component: Dict[str, Any], related_rows: List[Dict[str, Any]], exported_at: str) -> None:
    tables = _build_component_tables(component, related_rows)
    cache_path = _component_cache_path(component)
    manual_rows = _get_manual_rows(cache_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Info"
    ws.append(["Campo", "Valore"])
    info_rows = [
        ["Elemento", component.get("label", "")],
        ["Slug", component.get("slug", "")],
        ["Sezione", component.get("section", "")],
        ["Meta", component.get("meta", "")],
        ["Descrizione", component.get("description", "")],
        ["Report collegati", tables.get("related_count", 0)],
        ["Aggiornato il", exported_at],
        ["File Excel", cache_path.name],
        ["Link app", f"/MedicairGeek/infoImpianto/component/{component.get('slug', '')}"],
    ]
    for row in info_rows:
        ws.append(row)
    _autosize_worksheet(ws)

    network_rows = component.get("network_rows") or []
    if network_rows:
        _append_rows_sheet(
            wb,
            "IP impianto",
            ["Nome", "IP"],
            [[row.get("name", ""), row.get("ip", "")] for row in network_rows],
        )

    if component.get("show_qr_table"):
        _append_rows_sheet(
            wb,
            "QR sostituiti",
            ["Data", "Riferimento", "QR sostituiti", "Descrizione", "Link app"],
            [
                [
                    item.get("data", "-"),
                    item.get("riferimento", "-"),
                    item.get("sostituito_qr", "-"),
                    item.get("descrizione", "-"),
                    item.get("link", "#"),
                ]
                for item in tables.get("qr_changes", [])
            ],
        )

    _append_rows_sheet(
        wb,
        "Report",
        ["Data", "Titolo", "Tipo", "Descrizione", "Rimosso", "Link app"],
        [
            [
                item.get("data", "-"),
                item.get("titolo", "-"),
                item.get("tipo", "-"),
                item.get("descrizione", "-"),
                item.get("rimosso", "-"),
                item.get("link", "#"),
            ]
            for item in tables.get("reports", [])
        ],
    )
    _append_manual_sheet(wb, manual_rows)
    wb.save(cache_path)
    wb.close()


def _write_index_cache(rows: List[Dict[str, Any]], exported_at: str) -> None:
    manual_rows = _get_manual_rows(INDEX_CACHE_PATH)
    wb = Workbook()

    ws_meta = wb.active
    ws_meta.title = "Meta"
    ws_meta.append(["Campo", "Valore"])
    meta_rows = [
        ["Cache version", CACHE_VERSION],
        ["Aggiornato il", exported_at],
        ["Cartella cache", str(INFO_IMPIANTO_DIR)],
        ["File indice", INDEX_CACHE_PATH.name],
    ]
    for row in meta_rows:
        ws_meta.append(row)
    _autosize_worksheet(ws_meta)

    robot_sheet_rows: List[List[Any]] = []
    for robot in ROBOT_REGISTRY:
        related_rows = _get_related_rows_for_robot(robot, rows)
        robot_view = _decorate_robot(robot, related_rows)
        tables = _build_robot_tables(robot, related_rows)
        robot_sheet_rows.append(
            [
                robot_view.get("label", ""),
                robot_view.get("id", ""),
                robot_view.get("full_id", ""),
                robot_view.get("clean_ip", ""),
                "SI" if robot_view.get("is_removed") else "NO",
                robot_view.get("latest_report_label") or "-",
                tables.get("related_count", 0),
                _robot_cache_path(robot).name,
            ]
        )
    _append_rows_sheet(
        wb,
        "Robot",
        ["Robot", "ID breve", "ID completo", "IP", "Stato rimosso", "Ultimo report", "Report collegati", "File Excel"],
        robot_sheet_rows,
    )

    for section_label, section_items in (("Hardware", HARDWARE_ITEMS), ("Software", SOFTWARE_ITEMS)):
        component_sheet_rows: List[List[Any]] = []
        for item in section_items:
            component = _get_component(item.get("slug", "")) or dict(item)
            related_rows = _get_related_rows_for_component(component, rows)
            tables = _build_component_tables(component, related_rows)
            component_sheet_rows.append(
                [
                    component.get("label", ""),
                    component.get("slug", ""),
                    item.get("meta", component.get("meta", "")),
                    tables.get("related_count", 0),
                    _component_cache_path(component).name,
                ]
            )
        _append_rows_sheet(
            wb,
            section_label,
            ["Elemento", "Slug", "Meta", "Report collegati", "File Excel"],
            component_sheet_rows,
        )

    _append_manual_sheet(wb, manual_rows)
    wb.save(INDEX_CACHE_PATH)
    wb.close()


def _write_manifest(exported_at: str, reason: str) -> None:
    payload = {
        "cache_version": CACHE_VERSION,
        "updated_at": exported_at,
        "reason": reason,
        "index_file": INDEX_CACHE_PATH.name,
        "cache_dir": str(INFO_IMPIANTO_DIR),
        "robot_files": [_robot_cache_path(robot).name for robot in ROBOT_REGISTRY],
        "component_files": [_component_cache_path(component).name for component in COMPONENT_REGISTRY],
    }
    CACHE_MANIFEST_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _cache_missing() -> bool:
    if not INDEX_CACHE_PATH.exists():
        return True
    if any(not _robot_cache_path(robot).exists() for robot in ROBOT_REGISTRY):
        return True
    if any(not _component_cache_path(component).exists() for component in COMPONENT_REGISTRY):
        return True
    return False


def refresh_info_impianto_cache(
    rows: Optional[List[Dict[str, Any]]] = None,
    robot_ids: Optional[set[str]] = None,
    component_slugs: Optional[set[str]] = None,
    *,
    force_full: bool = False,
    reason: str = "manual",
) -> None:
    _ensure_cache_dir()
    exported_at = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    base_rows = rows if rows is not None else _read_excel_rows(limit=5000)

    with CACHE_LOCK:
        target_robots = ROBOT_REGISTRY
        target_components = COMPONENT_REGISTRY

        if not force_full and robot_ids is not None:
            normalized_ids = {_normalize_str(value) for value in robot_ids if _normalize_str(value)}
            target_robots = [
                robot
                for robot in ROBOT_REGISTRY
                if robot.get("id") in normalized_ids or robot.get("full_id") in normalized_ids
            ]

        if not force_full and component_slugs is not None:
            normalized_slugs = {_normalize_str(value).lower() for value in component_slugs if _normalize_str(value)}
            target_components = [
                component
                for component in COMPONENT_REGISTRY
                if _normalize_str(component.get("slug")).lower() in normalized_slugs
            ]

        for robot in target_robots:
            related_rows = _get_related_rows_for_robot(robot, base_rows)
            _write_robot_cache(robot, related_rows, exported_at)

        for component in target_components:
            related_rows = _get_related_rows_for_component(component, base_rows)
            _write_component_cache(component, related_rows, exported_at)

        _write_index_cache(base_rows, exported_at)
        _write_manifest(exported_at, reason)


def _extract_targets_from_row(row: Optional[Dict[str, Any]]) -> Tuple[set[str], set[str]]:
    robot_ids: set[str] = set()
    component_slugs: set[str] = set()

    if not row:
        return robot_ids, component_slugs

    for robot in ROBOT_REGISTRY:
        if _robot_matches(row.get("robot"), robot):
            robot_ids.add(robot.get("id", ""))
            if robot.get("full_id"):
                robot_ids.add(robot.get("full_id", ""))

    for component in COMPONENT_REGISTRY:
        if _component_matches(row.get("robot"), component):
            component_slugs.add(_normalize_str(component.get("slug")).lower())

    return robot_ids, component_slugs


def refresh_info_impianto_cache_for_report(report_id: int, previous_report: Optional[Dict[str, Any]] = None) -> None:
    base_rows = _read_excel_rows(limit=5000)
    current_report = next((row for row in base_rows if int(row.get("id", -1)) == int(report_id)), None)

    robot_ids: set[str] = set()
    component_slugs: set[str] = set()

    for row in (previous_report, current_report):
        row_robot_ids, row_component_slugs = _extract_targets_from_row(row)
        robot_ids.update(row_robot_ids)
        component_slugs.update(row_component_slugs)

    refresh_info_impianto_cache(
        rows=base_rows,
        robot_ids=robot_ids,
        component_slugs=component_slugs,
        force_full=_cache_missing(),
        reason=f"report:{report_id}",
    )


def ensure_info_impianto_cache(force: bool = False) -> None:
    if force or _cache_missing():
        refresh_info_impianto_cache(force_full=True, reason="startup")


def _read_home_cache() -> Optional[Dict[str, Any]]:
    ensure_info_impianto_cache()
    if not INDEX_CACHE_PATH.exists():
        return None

    wb = load_workbook(INDEX_CACHE_PATH, data_only=True)
    try:
        robot_rows = _sheet_to_dict_rows(wb["Robot"]) if "Robot" in wb.sheetnames else []
        hardware_rows = _sheet_to_dict_rows(wb["Hardware"]) if "Hardware" in wb.sheetnames else []
        software_rows = _sheet_to_dict_rows(wb["Software"]) if "Software" in wb.sheetnames else []
    finally:
        wb.close()

    robots: List[Dict[str, Any]] = []
    for row in robot_rows:
        robot_id = row.get("ID breve", "")
        full_id = row.get("ID completo", "")
        is_removed = _is_yes_value(row.get("Stato rimosso", ""))
        robots.append(
            {
                "id": robot_id,
                "full_id": full_id,
                "label": row.get("Robot", f"Robot {robot_id}"),
                "clean_ip": _clean_ip(row.get("IP", "")),
                "display_id": f"{robot_id} ({full_id})" if full_id else robot_id,
                "is_removed": is_removed,
                "status_label": "Robot rimosso = SI" if is_removed else "Robot rimosso = NO",
                "latest_report_label": row.get("Ultimo report", ""),
                "related_count": _safe_int(row.get("Report collegati", 0)),
                "cache_file": row.get("File Excel", ""),
            }
        )

    def _map_component_rows(rows: List[Dict[str, str]], defaults: List[Dict[str, str]]) -> List[Dict[str, Any]]:
        default_map = {item.get("slug", ""): item for item in defaults}
        items: List[Dict[str, Any]] = []
        for row in rows:
            slug = row.get("Slug", "")
            default_item = default_map.get(slug, {})
            items.append(
                {
                    "label": row.get("Elemento", default_item.get("label", slug)),
                    "slug": slug,
                    "meta": row.get("Meta", default_item.get("meta", "")),
                    "related_count": _safe_int(row.get("Report collegati", 0)),
                    "cache_file": row.get("File Excel", ""),
                }
            )
        return items or defaults

    return {
        "robots": robots or _build_robot_list(),
        "hardware_items": _map_component_rows(hardware_rows, HARDWARE_ITEMS),
        "software_items": _map_component_rows(software_rows, SOFTWARE_ITEMS),
    }


def _read_robot_cache(robot: Dict[str, str]) -> Optional[Dict[str, Any]]:
    ensure_info_impianto_cache()
    cache_path = _robot_cache_path(robot)
    if not cache_path.exists():
        return None

    wb = load_workbook(cache_path, data_only=True)
    try:
        info_map = _sheet_to_key_value_map(wb["Info"] if "Info" in wb.sheetnames else wb.active)
        replaced_rows = _sheet_to_dict_rows(wb["Pezzi sostituiti"]) if "Pezzi sostituiti" in wb.sheetnames else []
        manut_rows = _sheet_to_dict_rows(wb["Manutenzioni"]) if "Manutenzioni" in wb.sheetnames else []
        incident_rows = _sheet_to_dict_rows(wb["Incidenti"]) if "Incidenti" in wb.sheetnames else []
    finally:
        wb.close()

    robot_id = info_map.get("ID breve", robot.get("id", ""))
    full_id = info_map.get("ID completo", robot.get("full_id", ""))
    is_removed = _is_yes_value(info_map.get("Stato rimosso", ""))

    robot_view = {
        "id": robot_id,
        "full_id": full_id,
        "label": info_map.get("Robot", robot.get("label", f"Robot {robot_id}")),
        "display_id": info_map.get("ID display", f"{robot_id} ({full_id})" if full_id else robot_id),
        "clean_ip": _clean_ip(info_map.get("IP", robot.get("ip", ""))),
        "is_removed": is_removed,
        "status_label": info_map.get("Status label", "Robot rimosso = SI" if is_removed else "Robot rimosso = NO"),
        "latest_report_label": info_map.get("Ultimo report", ""),
        "cache_file": info_map.get("File Excel", cache_path.name),
    }

    replaced_parts = [
        {
            "data": row.get("Data", "-"),
            "pezzo": row.get("Pezzo", "-"),
            "link": row.get("Link app", "#"),
        }
        for row in replaced_rows
    ]
    manutenzioni = [
        {
            "data": row.get("Data", "-"),
            "titolo": row.get("Titolo", "-"),
            "parti_coinvolte": row.get("Parti coinvolte", "-"),
            "descrizione": row.get("Descrizione", "-"),
            "link": row.get("Link app", "#"),
        }
        for row in manut_rows
    ]
    incidenti = [
        {
            "data": row.get("Data", "-"),
            "titolo": row.get("Titolo", "-"),
            "tipo": row.get("Tipo", "-"),
            "descrizione": row.get("Descrizione", "-"),
            "rimosso": row.get("Rimosso", "-"),
            "link": row.get("Link app", "#"),
        }
        for row in incident_rows
    ]

    return {
        "robot": robot_view,
        "replaced_parts": replaced_parts,
        "manutenzioni": manutenzioni,
        "incidenti": incidenti,
        "related_count": _safe_int(info_map.get("Report collegati", 0)),
    }


def _read_component_cache(component: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    ensure_info_impianto_cache()
    cache_path = _component_cache_path(component)
    if not cache_path.exists():
        return None

    wb = load_workbook(cache_path, data_only=True)
    try:
        info_map = _sheet_to_key_value_map(wb["Info"] if "Info" in wb.sheetnames else wb.active)
        report_rows = _sheet_to_dict_rows(wb["Report"]) if "Report" in wb.sheetnames else []
        qr_rows = _sheet_to_dict_rows(wb["QR sostituiti"]) if "QR sostituiti" in wb.sheetnames else []
        network_rows = _sheet_to_dict_rows(wb["IP impianto"]) if "IP impianto" in wb.sheetnames else []
    finally:
        wb.close()

    component_view = dict(component)
    component_view["label"] = info_map.get("Elemento", component.get("label", ""))
    component_view["meta"] = info_map.get("Meta", component.get("meta", ""))
    component_view["description"] = info_map.get("Descrizione", component.get("description", ""))
    component_view["cache_file"] = info_map.get("File Excel", cache_path.name)

    reports = [
        {
            "data": row.get("Data", "-"),
            "titolo": row.get("Titolo", "-"),
            "tipo": row.get("Tipo", "-"),
            "descrizione": row.get("Descrizione", "-"),
            "rimosso": row.get("Rimosso", "-"),
            "link": row.get("Link app", "#"),
        }
        for row in report_rows
    ]
    qr_changes = [
        {
            "data": row.get("Data", "-"),
            "riferimento": row.get("Riferimento", "-"),
            "sostituito_qr": row.get("QR sostituiti", "-"),
            "descrizione": row.get("Descrizione", "-"),
            "link": row.get("Link app", "#"),
        }
        for row in qr_rows
    ]
    network_ip_rows = [{"name": row.get("Nome", ""), "ip": row.get("IP", "")} for row in network_rows]

    return {
        "component": component_view,
        "reports": reports,
        "qr_changes": qr_changes,
        "related_count": _safe_int(info_map.get("Report collegati", 0)),
        "network_rows": network_ip_rows or component.get("network_rows", []),
    }


@info_impianto_bp.get("/MedicairGeek/infoImpianto")
def info_impianto_home():
    active_tab = _normalize_str(request.args.get("tab", "robot")).lower() or "robot"
    if active_tab not in {"robot", "hardware", "software"}:
        active_tab = "robot"

    cached_home = _read_home_cache() or {}

    return render_template(
        "infoImpianto.html",
        title="Info Impianto",
        active_tab=active_tab,
        robots=cached_home.get("robots", _build_robot_list()),
        hardware_items=cached_home.get("hardware_items", HARDWARE_ITEMS),
        software_items=cached_home.get("software_items", SOFTWARE_ITEMS),
    )


@info_impianto_bp.get("/MedicairGeek/infoImpianto/robot/<robot_id>")
def info_impianto_robot_detail(robot_id: str):
    robot = _get_robot(robot_id)
    if not robot:
        abort(404)

    cached_robot = _read_robot_cache(robot)
    if cached_robot:
        return render_template(
            "infoImpiantoRobotDetail.html",
            title=f"Dettaglio {cached_robot['robot']['label']}",
            robot=cached_robot["robot"],
            replaced_parts=cached_robot["replaced_parts"],
            manutenzioni=cached_robot["manutenzioni"],
            incidenti=cached_robot["incidenti"],
            related_count=cached_robot["related_count"],
        )

    related_rows = _get_related_rows_for_robot(robot)
    robot_view = _decorate_robot(robot, related_rows)
    tables = _build_robot_tables(robot, related_rows)

    return render_template(
        "infoImpiantoRobotDetail.html",
        title=f"Dettaglio {robot_view['label']}",
        robot=robot_view,
        replaced_parts=tables["replaced_parts"],
        manutenzioni=tables["manutenzioni"],
        incidenti=tables["incidenti"],
        related_count=tables["related_count"],
    )


@info_impianto_bp.get("/MedicairGeek/infoImpianto/component/<slug>")
def info_impianto_component_detail(slug: str):
    component = _get_component(slug)
    if not component:
        abort(404)

    cached_component = _read_component_cache(component)
    if cached_component:
        component_view = cached_component["component"]
        return render_template(
            "infoImpiantoComponentDetail.html",
            title=f"Dettaglio {component_view['label']}",
            component=component_view,
            reports=cached_component["reports"],
            qr_changes=cached_component["qr_changes"],
            related_count=cached_component["related_count"],
            show_qr_table=bool(component_view.get("show_qr_table")),
            network_rows=cached_component.get("network_rows", []),
            section_label=("Hardware" if component_view.get("section") == "hardware" else "Software"),
        )

    related_rows = _get_related_rows_for_component(component)
    tables = _build_component_tables(component, related_rows)

    return render_template(
        "infoImpiantoComponentDetail.html",
        title=f"Dettaglio {component['label']}",
        component=component,
        reports=tables["reports"],
        qr_changes=tables["qr_changes"],
        related_count=tables["related_count"],
        show_qr_table=bool(component.get("show_qr_table")),
        network_rows=component.get("network_rows", []),
        section_label=("Hardware" if component.get("section") == "hardware" else "Software"),
    )


@info_impianto_bp.get("/MedicairGeek/infoImpianto/placeholder/<section>/<slug>")
def info_impianto_placeholder(section: str, slug: str):
    label = _normalize_str(request.args.get("label")) or slug.replace("-", " ").title()
    section_label = {
        "hardware": "Hardware",
        "software": "Software",
        "robot": "Robot",
    }.get(section.lower(), "Info Impianto")

    return render_template(
        "infoImpiantoPlaceholder.html",
        title=f"{label} - Pagina in costruzione",
        section_label=section_label,
        item_label=label,
    )
