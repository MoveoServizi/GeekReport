from __future__ import annotations

import shutil
import threading
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any

from flask import Blueprint, jsonify, render_template, request, abort
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

from config import DESTINATARI,DESTINATARI_ALL_REPORT, REPORT_BASE_DIR
from email_sender import EmailSender
from modelli_latex import cleanup_latex_tmp, crea_report


# =====================
# Config / Costanti
# =====================

APP_TITLE = "Report Medicair - Incidenti Robot"
BASE_DIR = Path(__file__).resolve().parent

REPORT_DIR = REPORT_BASE_DIR
EXCEL_PATH = REPORT_DIR / "Incidenti_robot.xlsx"

ALLOWED_EXT = {
    "jpg", "jpeg", "png", "gif", "webp", "heic", "bmp", "tiff",
    "mp4", "mov", "m4v", "avi", "mkv", "webm",
    "pdf", "doc", "docx", "xls", "xlsx",
}

MAX_TITOLO_LEN = 30

CATEGORIE_LIST = [
    "Incidente",
    "Incidente Grave",
    "Problema Software",
    "Problema Hardware",
    "Intervento Manutenzione",
    "Altro",
]

ROBOT_LIST = [
    "1216278", "1216279", "1216292", "1216294", "1216302", "1216306", "1216313", "1216314",
    "1216325", "1216337", "1216339", "1216340", "1216348", "1216349", "1216350",
    "Tutti", "Pavimento", "WorkingStation1", "WorkingStation2", "WorkingStation3",
    "ChargingStation", "Altro", "Scaffale",
]

ZONA_LIST = [
    "Avvio Robot",
    "Corridoio principale",
    "Corridoi",
    "Station 1",
    "Station 2",
    "station 3",
    "Manutenzione",
    "Altro",
]

LUCI_CAMPO1 = ["Fisse", "Lampeggianti", "Non applicabile"]
LUCI_CAMPO2 = ["Bianca", "Rossa", "Verde", "Blu", "Gialla", "Viola", "Altro"]

HEADERS = [
    "id",
    "data",
    "ora",
    "Categoria",
    "Titolo",
    "robot",
    "scaffale",
    "cella",
    "zona",
    "luci robot",
    "errore",
    "note",
    "rimosso",
    "risoluzione",
    "redatto_da", 
    "data_update1",
    "update1",
    "data_update2",
    "update2",
]


# =====================
# Job store (in-memory)
# =====================

JOBS: Dict[str, Dict[str, Any]] = {}
JOBS_LOCK = threading.Lock()
JOB_TTL_SECONDS = 6 * 60 * 60  # 6 ore


def _job_set(job_id: str, **kwargs) -> None:
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return
        job.update(kwargs)


def _job_get(job_id: str) -> Optional[Dict[str, Any]]:
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        return dict(job) if job else None


def _jobs_gc() -> None:
    now = time.time()
    to_delete = []

    with JOBS_LOCK:
        for jid, job in JOBS.items():
            created = job.get("created_ts", now)
            if now - created > JOB_TTL_SECONDS:
                to_delete.append(jid)

        for jid in to_delete:
            del JOBS[jid]


# =====================
# Helpers generici
# =====================

def normalize_spaces(value: str) -> str:
    return " ".join((value or "").strip().split())


def sanitize_titolo(value: str) -> str:
    return normalize_spaces(value)[:MAX_TITOLO_LEN]


def allowed_file(filename: str) -> bool:
    if not filename or "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in ALLOWED_EXT


def safe_cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_bool_like_si_no(value: str, default: str = "no") -> str:
    v = normalize_spaces(value).lower()
    if v in ("si", "sì", "yes", "y", "true", "1"):
        return "si"
    if v in ("no", "n", "false", "0"):
        return "no"
    return default


def parse_datetime_local(value: str) -> datetime:
    return datetime.strptime(value, "%Y-%m-%dT%H:%M")


def format_date_it(dt: datetime) -> str:
    return dt.strftime("%d/%m/%Y")


def format_time_it(dt: datetime) -> str:
    return dt.strftime("%H:%M")


def format_datetime_it(dt: datetime) -> str:
    return dt.strftime("%d/%m/%Y %H:%M")


# =====================
# Excel helpers
# =====================

def ensure_report_assets() -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    if not EXCEL_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Incidenti"
        ws.append(HEADERS)
        wb.save(EXCEL_PATH)
        wb.close()
    else:
        ensure_excel_headers()


def ensure_excel_headers() -> None:
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    current_headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    changed = False

    for idx, header in enumerate(HEADERS, start=1):
        current_val = current_headers[idx - 1] if idx - 1 < len(current_headers) else None
        if safe_cell_str(current_val) != header:
            ws.cell(row=1, column=idx, value=header)
            changed = True

    if changed:
        wb.save(EXCEL_PATH)

    wb.close()


def get_header_index_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[str(value).strip()] = col
    return headers


def get_next_id() -> int:
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    last_id = 0
    for row in range(ws.max_row, 1, -1):
        val = ws.cell(row=row, column=1).value
        if val is not None and str(val).strip() != "":
            try:
                last_id = int(val)
                break
            except Exception:
                pass

    wb.close()
    return last_id + 1


def append_row(row_values: list[Any]) -> None:
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.append(row_values)
    wb.save(EXCEL_PATH)
    wb.close()


def find_report_row_by_id(report_id: int) -> Optional[int]:
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    row_found = None
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        try:
            if int(val) == int(report_id):
                row_found = row
                break
        except Exception:
            pass

    wb.close()
    return row_found


def get_report_by_id(report_id: int) -> Optional[dict[str, Any]]:
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = get_header_index_map(ws)

    row_found = None
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        try:
            if int(val) == int(report_id):
                row_found = row
                break
        except Exception:
            pass

    if row_found is None:
        wb.close()
        return None

    data: dict[str, Any] = {}
    for key, col in headers.items():
        data[key] = ws.cell(row=row_found, column=col).value

    wb.close()

    update1 = safe_cell_str(data.get("update1"))
    update2 = safe_cell_str(data.get("update2"))
    data["has_update"] = bool(update1 or update2)

    return data


def update_report_fields(report_id: int, fields: dict[str, Any]) -> bool:
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = get_header_index_map(ws)

    row_found = None
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        try:
            if int(val) == int(report_id):
                row_found = row
                break
        except Exception:
            pass

    if row_found is None:
        wb.close()
        return False

    for key, value in fields.items():
        if key in headers:
            ws.cell(row=row_found, column=headers[key], value=value)

    wb.save(EXCEL_PATH)
    wb.close()
    return True


def add_report_update(report_id: int, update_text: str, update_dt: Optional[datetime] = None) -> dict[str, Any]:
    update_text = (update_text or "").strip()
    if not update_text:
        return {"ok": False, "error": "Testo update vuoto."}

    report = get_report_by_id(report_id)
    if not report:
        return {"ok": False, "error": "Report non trovato."}

    update_dt = update_dt or datetime.now()
    upd1 = safe_cell_str(report.get("update1"))
    upd2 = safe_cell_str(report.get("update2"))

    if not upd1:
        ok = update_report_fields(
            report_id,
            {
                "data_update1": format_datetime_it(update_dt),
                "update1": update_text,
            },
        )
        return {
            "ok": ok,
            "slot": 1 if ok else None,
            "error": None if ok else "Errore salvataggio update 1.",
        }

    if not upd2:
        ok = update_report_fields(
            report_id,
            {
                "data_update2": format_datetime_it(update_dt),
                "update2": update_text,
            },
        )
        return {
            "ok": ok,
            "slot": 2 if ok else None,
            "error": None if ok else "Errore salvataggio update 2.",
        }

    return {"ok": False, "error": "Questo report ha già 2 update."}

def find_report_folder(report_id: int) -> Optional[Path]:
    if not REPORT_DIR.exists():
        return None

    prefix = f"{report_id}_"
    candidates = [p for p in REPORT_DIR.iterdir() if p.is_dir() and p.name.startswith(prefix)]
    if not candidates:
        return None

    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def build_allegati_list_tex(folder_path: Path) -> str:
    files = []
    for p in sorted(folder_path.iterdir(), key=lambda x: x.name.lower()):
        if not p.is_file():
            continue
        if p.suffix.lower() == ".pdf" and p.name.upper().startswith("REPORT_"):
            continue
        files.append(p.name)

    if not files:
        return r"\FileItem{Nessun allegato}"

    return "\n".join([rf"\FileItem{{{name}}}" for name in files])


def regenerate_report_pdf(report_id: int) -> dict[str, Any]:
    try:
        report = get_report_by_id(report_id)
        if not report:
            return {"ok": False, "message": "Report non trovato per rigenerazione PDF."}

        folder_path = find_report_folder(report_id)
        if not folder_path:
            return {"ok": False, "message": "Cartella report non trovata per rigenerazione PDF."}

        data_str = safe_cell_str(report.get("data"))
        ora_str = safe_cell_str(report.get("ora"))
        titolo = safe_cell_str(report.get("Titolo"))
        categoria = safe_cell_str(report.get("Categoria"))
        robot = safe_cell_str(report.get("robot"))
        scaffale = safe_cell_str(report.get("scaffale")) or "senza scaffale"
        cella = safe_cell_str(report.get("cella"))
        zona = safe_cell_str(report.get("zona"))
        luci_robot = safe_cell_str(report.get("luci robot"))
        errore = safe_cell_str(report.get("errore"))
        rimosso = safe_cell_str(report.get("rimosso"))
        risoluzione = safe_cell_str(report.get("risoluzione"))
        descrizione = safe_cell_str(report.get("note"))
        data_update1 = safe_cell_str(report.get("data_update1"))
        update1 = safe_cell_str(report.get("update1"))
        data_update2 = safe_cell_str(report.get("data_update2"))
        update2 = safe_cell_str(report.get("update2"))

        allegati_list_tex = build_allegati_list_tex(folder_path)

       

        campi_report = {
            "ReportID": str(report_id),
            "DataIncidente": data_str,
            "OraIncidente": ora_str,
            "Titolo": titolo,
            "Categoria": categoria,
            "Robot": robot,
            "Scaffale": scaffale,
            "Cella": cella,
            "Zona": zona,
            "LuciRobot": luci_robot,
            "Errore": errore,
            "Rimosso": rimosso,
            "Risoluzione": risoluzione,
            "Descrizione": descrizione,
            "DataUpdate1": data_update1,
            "Update1": update1,
            "DataUpdate2": data_update2,
            "Update2": update2,
            "AllegatiList": allegati_list_tex,
        }

        nome_pdf = None
        existing_pdf = None
        for p in folder_path.iterdir():
            if p.is_file() and p.suffix.lower() == ".pdf" and p.name.upper().startswith("REPORT_"):
                existing_pdf = p
                nome_pdf = p.stem
                break

        if not nome_pdf:
            nome_pdf = f"REPORT_{report_id}"

        res = crea_report("modello", campi_report, nome_file=nome_pdf)
        output_pdf = folder_path / f"{nome_pdf}.pdf"
        shutil.copy2(res.pdf_path, output_pdf)

        try:
            cleanup_latex_tmp()
        except Exception:
            pass

        return {
            "ok": True,
            "message": f"PDF rigenerato: {output_pdf.name}"
        }

    except Exception as e:
        return {
            "ok": False,
            "message": f"Errore rigenerazione PDF: {e}"
        }

def find_report_folder(report_id: int) -> Optional[Path]:
    if not REPORT_DIR.exists():
        return None

    prefix = f"{report_id}_"
    candidates = [p for p in REPORT_DIR.iterdir() if p.is_dir() and p.name.startswith(prefix)]
    if not candidates:
        return None

    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def save_uploaded_files_to_report_folder(report_id: int, files) -> dict[str, Any]:
    folder_path = find_report_folder(report_id)
    if not folder_path:
        return {"ok": False, "error": "Cartella report non trovata.", "saved_files": [], "warnings": []}

    saved_files: list[str] = []
    warnings: list[str] = []

    for f in files:
        if not f or not f.filename:
            continue

        if not allowed_file(f.filename):
            warnings.append(f"File ignorato: {f.filename}")
            continue

        safe = secure_filename(f.filename)
        if not safe:
            warnings.append(f"Nome file non valido: {f.filename}")
            continue

        dest = folder_path / safe
        i = 1
        while dest.exists():
            dest = folder_path / f"{dest.stem}_{i}{dest.suffix}"
            i += 1

        try:
            f.save(dest)
            saved_files.append(dest.name)
        except Exception as e:
            warnings.append(f"Errore salvataggio {f.filename}: {e}")

    return {
        "ok": True,
        "saved_files": saved_files,
        "warnings": warnings,
        "folder_path": folder_path,
    }
# =====================
# Email
# =====================

def _send_report_email(
    report_id: int,
    dt: datetime,
    titolo: str,
    categoria: str,
    robots: list[str],
    note: str,
    destinatari: list[str],
    allegati: list[Path],
) -> dict[str, Any]:
    if not destinatari:
        return {"ok": True, "sent": 0, "errors": []}

    sender = EmailSender()
    template_name = "REPORT INCIDENTE"

    fields = {
        "data": format_datetime_it(dt),
        "robots": ", ".join(robots),
        "note": (note or "").strip(),
        "titolo": titolo,
        "categoria": categoria,
        "report_id": str(report_id),
    }

    attachments = [str(p) for p in (allegati or []) if p and p.exists() and p.is_file()]

    sent = 0
    errors: list[str] = []

    for dest in destinatari:
        try:
            res = sender.send_template(dest, template_name, fields, attachments)
            if getattr(res, "ok", False):
                sent += 1
            else:
                errors.append(f"{dest}: {getattr(res, 'error', 'Errore sconosciuto')}")
        except Exception as e:
            errors.append(f"{dest}: {e}")

    return {
        "ok": sent > 0 and not errors,
        "sent": sent,
        "errors": errors,
    }


# =====================
# Worker
# =====================

def _run_job(job_id: str, payload: Dict[str, Any]) -> None:
    """
    Worker che esegue:
    - Excel
    - PDF
    - Email
    - Cleanup latex/tmp solo se non ci sono errori
    """
    worker_errors: list[str] = []
    pdf_report_path: Optional[Path] = None

    try:
        ensure_report_assets()

        dt: datetime = payload["dt"]
        dt_local: str = payload["dt_local"]
        titolo: str = payload["titolo"]
        categoria: str = payload["categoria"]
        robots: list[str] = payload["robots"]
        scaffale: str = payload["scaffale"]
        cella: str = payload["cella"]
        zona: str = payload["zona"]
        errore: str = payload["errore"]
        descrizione: str = payload["descrizione"]
        luci_robot: str = payload["luci_robot"]
        rimosso: str = payload["rimosso"]
        risoluzione: str = payload["risoluzione"]
        redatto_da: str = payload["redatto_da"]
        next_id: int = payload["next_id"]
        folder_name: str = payload["folder_name"]
        folder_path: Path = payload["folder_path"]
        saved_file_paths: list[Path] = payload["saved_file_paths"]

        saved_files = [p.name for p in saved_file_paths]

        # ---- EXCEL
        _job_set(job_id, phase="EXCEL", percent=45, message="Scrittura registro Excel…")

        row = [
            next_id,
            format_date_it(dt),
            format_time_it(dt),
            categoria,
            titolo,
            ", ".join(robots),
            scaffale if scaffale else "senza scaffale",
            cella,
            zona,
            luci_robot,
            errore,
            descrizione,
            rimosso,
            risoluzione,
            redatto_da,
            "",
            "",
            "",
            "",
        ]
        append_row(row)

        # ---- PDF
        _job_set(job_id, phase="PDF", percent=62, message="Generazione PDF (LaTeX)…")

        try:
            if saved_file_paths:
                allegati_list_tex = "\n".join([rf"\FileItem{{{p.name}}}" for p in saved_file_paths])
            else:
                allegati_list_tex = r"\FileItem{Nessun allegato}"

            campi_report = {
                "ReportID": str(next_id),
                "DataIncidente": format_date_it(dt),
                "OraIncidente": format_time_it(dt),
                "Titolo": titolo,
                "Categoria": categoria,
                "Robot": ", ".join(robots),
                "Scaffale": scaffale if scaffale else "senza scaffale",
                "Cella": cella,
                "Zona": zona,
                "LuciRobot": luci_robot,
                "Errore": errore,
                "Rimosso": rimosso,
                "Risoluzione": risoluzione,
                "RedattoDa": redatto_da,
                "Descrizione": descrizione,
                "AllegatiList": allegati_list_tex,
            }

            nome_pdf = f"REPORT_{next_id}_{dt.strftime('%d-%m-%Y_%H-%M')}"
            res = crea_report("modello", campi_report, nome_file=nome_pdf)

            pdf_report_path = folder_path / res.pdf_path.name
            shutil.copy2(res.pdf_path, pdf_report_path)

            saved_files.append(pdf_report_path.name)
            saved_file_paths.append(pdf_report_path)

        except Exception as e:
            worker_errors.append(f"PDF non generato: {e}")
            _job_set(job_id, message=f"PDF non generato: {e}")

        # ---- EMAIL
        _job_set(job_id, phase="EMAIL", percent=80, message="Invio email…")

        destinari_email = list(DESTINATARI)
        if categoria == "Incidente":
            destinari_email = list(DESTINATARI_ALL_REPORT)
        else:
            destinari_email = list(DESTINATARI)

        try:
            email_res = _send_report_email(
                report_id=next_id,
                dt=dt,
                titolo=titolo,
                categoria=categoria,
                robots=robots,
                note=descrizione,
                destinatari=destinari_email,
                allegati=list(saved_file_paths),
            )
            if email_res.get("errors"):
                worker_errors.extend([str(x) for x in email_res["errors"]])
        except Exception as e:
            email_res = {"ok": False, "sent": 0, "errors": [str(e)]}
            worker_errors.append(f"Errore email: {e}")

        # ---- CLEANUP solo se non ci sono stati errori
        cleanup_done = False
        if not worker_errors:
            _job_set(job_id, phase="CLEANUP", percent=92, message="Pulizia file temporanei…")
            try:
                cleanup_latex_tmp()
                cleanup_done = True
            except Exception as e:
                worker_errors.append(f"Pulizia temp fallita: {e}")

        # ---- DONE
        _job_set(
            job_id,
            phase="DONE",
            percent=100,
            message="Completato." if not worker_errors else "Completato con avvisi.",
            done=True,
            result={
                "report_id": next_id,
                "folder_name": folder_name,
                "saved_files": saved_files,
                "pdf_created": bool(pdf_report_path and pdf_report_path.exists()),
                "email": email_res,
                "dt_local": dt_local,
                "titolo": titolo,
                "categoria": categoria,
                "robots": robots,
                "cleanup_done": cleanup_done,
                "warnings": worker_errors,
                "has_attachments": bool(payload["saved_file_paths"]),
            },
        )

    except Exception as e:
        _job_set(
            job_id,
            phase="ERROR",
            done=True,
            error=str(e),
            message="Errore durante l'elaborazione.",
            percent=100,
        )


# =====================
# Blueprint
# =====================

report_incidente_bp = Blueprint("report_incidente", __name__)


@report_incidente_bp.get("/MedicairGeek/reportIncidente")
def report_form():
    _jobs_gc()
    now = datetime.now()
    dt_local = now.strftime("%Y-%m-%dT%H:%M")

    return render_template(
        "reportIncidente.html",
        title=APP_TITLE,
        dt_local=dt_local,
        robot_list=ROBOT_LIST,
        zona_list=ZONA_LIST,
        luci_campo1=LUCI_CAMPO1,
        luci_campo2=LUCI_CAMPO2,
        categorie_list=CATEGORIE_LIST,
        max_titolo_len=MAX_TITOLO_LEN,
    )


@report_incidente_bp.post("/MedicairGeek/reportIncidente/start")
def start_job():
    """
    Riceve form + file (multipart), salva gli upload nella cartella report,
    crea job e avvia thread worker.
    Gli allegati sono facoltativi.
    """
    _jobs_gc()
    ensure_report_assets()

    # ---- parse form
    dt_local = normalize_spaces(request.form.get("dt_local", ""))
    titolo_raw = request.form.get("titolo", "") or ""
    titolo = sanitize_titolo(titolo_raw)
    categoria = normalize_spaces(request.form.get("categoria", ""))
    robots = request.form.getlist("robots")
    scaffale = normalize_spaces(request.form.get("scaffale", "senza scaffale"))
    zona = normalize_spaces(request.form.get("zona", ""))
    errore = normalize_spaces(request.form.get("errore", ""))
    descrizione = (request.form.get("descrizione", "") or "").strip()
    luci_c1 = normalize_spaces(request.form.get("luci_c1", ""))
    luci_c2 = normalize_spaces(request.form.get("luci_c2", ""))
    cella = normalize_spaces(request.form.get("cella", ""))
    rimosso = parse_bool_like_si_no(request.form.get("rimosso", "no"), default="no")
    risoluzione = normalize_spaces(request.form.get("risoluzione", ""))
    redatto_da = normalize_spaces(request.form.get("redatto_da", ""))
    files = request.files.getlist("media")

    # ---- validation
    errors: list[str] = []

    if not dt_local:
        errors.append("Data e ora sono obbligatori.")

    if not titolo_raw.strip():
        errors.append("Titolo è obbligatorio.")
    elif len(normalize_spaces(titolo_raw)) > MAX_TITOLO_LEN:
        errors.append(f"Titolo troppo lungo: massimo {MAX_TITOLO_LEN} caratteri.")

    if not categoria:
        errors.append("Categoria è obbligatoria.")
    elif categoria not in CATEGORIE_LIST:
        errors.append("Categoria non valida.")

    if not robots:
        errors.append("Seleziona almeno un robot.")

    if not zona:
        errors.append("Zona è obbligatoria.")

    if not luci_c1 or not luci_c2:
        errors.append("Luci robot è obbligatorio (seleziona entrambi).")

    if not descrizione.strip():
        errors.append("Descrizione è obbligatoria.")

    invalid = [r for r in robots if r not in ROBOT_LIST]
    if invalid:
        errors.append(f"Robot non validi: {', '.join(invalid)}")

    if zona and zona not in ZONA_LIST:
        errors.append("Zona non valida.")

    if luci_c1 and luci_c1 not in LUCI_CAMPO1:
        errors.append("Luci campo 1 non valide.")

    if luci_c2 and luci_c2 not in LUCI_CAMPO2:
        errors.append("Luci campo 2 non valide.")

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    # ---- dt parse
    try:
        dt = parse_datetime_local(dt_local)
    except Exception:
        return jsonify({"ok": False, "errors": ["Formato data/ora non valido."]}), 400

    # ---- create IDs + folder
    next_id = get_next_id()
    folder_stamp = dt.strftime("%d-%m-%Y_%H-%M")
    folder_name = f"{next_id}_{folder_stamp}"
    folder_path = REPORT_DIR / folder_name
    folder_path.mkdir(parents=True, exist_ok=True)

    # ---- create job
    job_id = uuid.uuid4().hex
    with JOBS_LOCK:
        JOBS[job_id] = {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "created_ts": time.time(),
            "phase": "UPLOAD_SAVING",
            "percent": 10,
            "message": "Salvataggio allegati…",
            "done": False,
            "error": None,
            "result": None,
        }

    # ---- save uploads (facoltativi)
    saved_file_paths: list[Path] = []
    upload_warnings: list[str] = []

    for f in files:
        if not f or not f.filename:
            continue

        if not allowed_file(f.filename):
            upload_warnings.append(f"File ignorato: {f.filename}")
            continue

        safe = secure_filename(f.filename)
        if not safe:
            upload_warnings.append(f"Nome file non valido: {f.filename}")
            continue

        dest = folder_path / safe
        i = 1
        while dest.exists():
            dest = folder_path / f"{dest.stem}_{i}{dest.suffix}"
            i += 1

        try:
            f.save(dest)
            saved_file_paths.append(dest)
        except Exception as e:
            upload_warnings.append(f"Errore salvataggio {f.filename}: {e}")

    luci_robot = f"{luci_c1} - {luci_c2}".strip()

    payload = {
        "dt": dt,
        "dt_local": dt_local,
        "titolo": titolo,
        "categoria": categoria,
        "robots": robots,
        "scaffale": scaffale,
        "cella": cella,
        "zona": zona,
        "errore": errore,
        "descrizione": descrizione.strip(),
        "luci_robot": luci_robot,
        "rimosso": rimosso,
        "risoluzione": risoluzione,
        "redatto_da": redatto_da,
        "next_id": next_id,
        "folder_name": folder_name,
        "folder_path": folder_path,
        "saved_file_paths": saved_file_paths,
    }

    msg = "Upload completato. Avvio elaborazione…"
    if not saved_file_paths:
        msg = "Nessun allegato caricato. Avvio elaborazione…"

    _job_set(
        job_id,
        phase="UPLOAD_SAVED",
        percent=35,
        message=msg,
        upload_warnings=upload_warnings,
    )

    t = threading.Thread(target=_run_job, args=(job_id, payload), daemon=True)
    t.start()

    return jsonify(
        {
            "ok": True,
            "job_id": job_id,
            "has_attachments": bool(saved_file_paths),
            "upload_warnings": upload_warnings,
        }
    ), 200


@report_incidente_bp.get("/MedicairGeek/reportIncidente/status/<job_id>")
def job_status(job_id: str):
    _jobs_gc()
    job = _job_get(job_id)
    if not job:
        return jsonify({"ok": False, "error": "Job non trovato o scaduto."}), 404
    return jsonify({"ok": True, "job": job}), 200


@report_incidente_bp.get("/MedicairGeek/reportIncidente/success/<job_id>")
def job_success(job_id: str):
    _jobs_gc()
    job = _job_get(job_id)
    if not job:
        abort(404)

    if not job.get("done") or job.get("phase") != "DONE" or not job.get("result"):
        abort(404)

    r = job["result"]
    return render_template(
        "send_report_success.html",
        title=APP_TITLE,
        report_id=r.get("report_id"),
        folder_name=r.get("folder_name"),
        saved_files=r.get("saved_files", []),
        destinatari=r.get("destinatari", []),
    )


# =====================
# API report esistenti
# =====================

@report_incidente_bp.get("/MedicairGeek/reportIncidente/<int:report_id>")
def get_report_detail(report_id: int):
    ensure_report_assets()
    report = get_report_by_id(report_id)
    if not report:
        return jsonify({"ok": False, "error": "Report non trovato."}), 404
    return jsonify({"ok": True, "report": report}), 200


@report_incidente_bp.post("/MedicairGeek/reportIncidente/<int:report_id>/update")
def insert_report_update(report_id: int):
    """
    Inserisce un update nel primo slot libero.
    Massimo 2 update.
    Permette anche di aggiungere allegati alla cartella esistente.
    Non invia email.
    """
    try:
        ensure_report_assets()

        if request.content_type and "multipart/form-data" in request.content_type.lower():
            update_text = (request.form.get("update") or "").strip()
            files = request.files.getlist("media")
        else:
            data = request.get_json(silent=True) or {}
            update_text = (data.get("update") or "").strip()
            files = []

        if not update_text:
            return jsonify({"ok": False, "error": "Update obbligatorio."}), 400

        res = add_report_update(report_id, update_text)

        if not res.get("ok"):
            return jsonify(res), 400

        upload_res = save_uploaded_files_to_report_folder(report_id, files)
        if not upload_res.get("ok"):
            return jsonify({"ok": False, "error": upload_res.get("error", "Errore upload file.")}), 400

        pdf_result = regenerate_report_pdf(report_id)
        report = get_report_by_id(report_id)

        return jsonify(
            {
                "ok": True,
                "slot": res.get("slot"),
                "report": report,
                "saved_files": upload_res.get("saved_files", []),
                "upload_warnings": upload_res.get("warnings", []),
                "pdf_regenerated": bool(pdf_result and pdf_result.get("ok")),
                "pdf_message": pdf_result.get("message") if isinstance(pdf_result, dict) else None,
            }
        ), 200

    except Exception as e:
        return jsonify({"ok": False, "error": f"Errore interno durante l'update: {e}"}), 500

@report_incidente_bp.post("/MedicairGeek/reportIncidente/<int:report_id>/edit")
def edit_report(report_id: int):
    """
    Modifica i campi del report esistente.
    Permette anche di aggiungere allegati alla cartella esistente.
    Non invia email.
    """
    try:
        ensure_report_assets()

        files = []

        if request.content_type and "multipart/form-data" in request.content_type.lower():
            raw_data = request.form.to_dict(flat=True)
            files = request.files.getlist("media")
        else:
            raw_data = request.get_json(silent=True) or {}

        if not isinstance(raw_data, dict):
            return jsonify({"ok": False, "error": "Payload non valido."}), 400

        regenerate_pdf = str(raw_data.pop("_regenerate_pdf", "true")).lower() in ("true", "1", "yes", "si")

        allowed_fields = {
            "data",
            "ora",
            "Categoria",
            "Titolo",
            "robot",
            "scaffale",
            "cella",
            "zona",
            "luci robot",
            "errore",
            "note",
            "rimosso",
            "risoluzione",
            "data_update1",
            "update1",
            "data_update2",
            "update2",
        }

        fields_to_update: dict[str, Any] = {}

        for k, v in raw_data.items():
            if k not in allowed_fields:
                continue

            if isinstance(v, str):
                v = v.strip()

            if k == "Titolo":
                if not str(v).strip():
                    return jsonify({"ok": False, "error": "Titolo obbligatorio."}), 400
                if len(normalize_spaces(str(v))) > MAX_TITOLO_LEN:
                    return jsonify({
                        "ok": False,
                        "error": f"Titolo troppo lungo: massimo {MAX_TITOLO_LEN} caratteri."
                    }), 400
                v = sanitize_titolo(str(v))

            elif k == "Categoria":
                if v not in CATEGORIE_LIST:
                    return jsonify({"ok": False, "error": "Categoria non valida."}), 400

            elif k == "zona":
                if v and v not in ZONA_LIST:
                    return jsonify({"ok": False, "error": "Zona non valida."}), 400

            elif k == "rimosso":
                v = parse_bool_like_si_no(str(v), default="no")

            fields_to_update[k] = v

        upload_res = save_uploaded_files_to_report_folder(report_id, files)
        if not upload_res.get("ok"):
            return jsonify({"ok": False, "error": upload_res.get("error", "Errore upload file.")}), 400

        if not fields_to_update and not upload_res.get("saved_files"):
            return jsonify({"ok": False, "error": "Nessun campo valido o file da aggiornare."}), 400

        if fields_to_update:
            ok = update_report_fields(report_id, fields_to_update)
            if not ok:
                return jsonify({"ok": False, "error": "Report non trovato."}), 404

        pdf_result = None
        if regenerate_pdf:
            pdf_result = regenerate_report_pdf(report_id)

        report = get_report_by_id(report_id)

        return jsonify(
            {
                "ok": True,
                "message": "Report aggiornato correttamente.",
                "report": report,
                "saved_files": upload_res.get("saved_files", []),
                "upload_warnings": upload_res.get("warnings", []),
                "pdf_regenerated": bool(pdf_result and pdf_result.get("ok")),
                "pdf_message": pdf_result.get("message") if isinstance(pdf_result, dict) else None,
            }
        ), 200

    except Exception as e:
        return jsonify({
            "ok": False,
            "error": f"Errore interno durante la modifica: {e}"
        }), 500
    
@report_incidente_bp.get("/MedicairGeek/reportIncidente/<int:report_id>/update-info")
def get_report_update_info(report_id: int):
    """
    Endpoint comodo per sapere se il report ha update e quanti slot liberi restano.
    """
    ensure_report_assets()
    report = get_report_by_id(report_id)
    if not report:
        return jsonify({"ok": False, "error": "Report non trovato."}), 404

    update1 = safe_cell_str(report.get("update1"))
    update2 = safe_cell_str(report.get("update2"))

    used_slots = 0
    if update1:
        used_slots += 1
    if update2:
        used_slots += 1

    return jsonify(
        {
            "ok": True,
            "has_update": bool(update1 or update2),
            "used_slots": used_slots,
            "free_slots": max(0, 2 - used_slots),
            "can_add_update": used_slots < 2,
            "data_update1": report.get("data_update1"),
            "update1": report.get("update1"),
            "data_update2": report.get("data_update2"),
            "update2": report.get("update2"),
        }
    ), 200



