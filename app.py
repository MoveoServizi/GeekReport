import os
import re
import shutil
from datetime import datetime
from pathlib import Path

from flask import Flask, render_template, request, redirect, url_for, flash
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from typing import Optional
# PDF LaTeX
from modelli_latex import crea_report, cleanup_latex_tmp

# Email
from email_sender import EmailSender
from email_config import DESTINATARI

APP_TITLE = "Report Medicair - Incidenti Robot"
BASE_DIR = Path(__file__).resolve().parent

REPORT_DIR = BASE_DIR / "report"
EXCEL_PATH = REPORT_DIR / "Incidenti_robot.xlsx"

# Upload
ALLOWED_EXT = {
    "jpg", "jpeg", "png", "gif", "webp", "heic",
    "mp4", "mov", "m4v", "avi", "mkv", "webm"
}
MAX_CONTENT_LENGTH = 300 * 1024 * 1024  # 300MB

ROBOT_LIST = [
    "16278", "16279", "16292", "16294", "16302", "16306", "16314",
    "16325", "16337", "16339", "16340", "16348", "16349", "16350"
]

ZONA_LIST = [
    "Avvio Robot",
    "Corridoio principale",
    "corridoi",
    "Station 1",
    "Station 2",
    "station 3",
    "ingresso Ws1",
    "ingresso WS1/2"
]

LUCI_CAMPO1 = ["Fisse", "Lampeggianti"]
LUCI_CAMPO2 = ["bianca", "blu", "verde", "rossa", "gialla", "altro"]

HEADERS = [
    "id", "data", "ora", "robot", "scaffale", "cella", "zona",
    "luci robot", "errore", "note", "rimosso", "risoluzione"
]


def ensure_report_assets():
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    if not EXCEL_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Incidenti"
        ws.append(HEADERS)
        wb.save(EXCEL_PATH)
        wb.close()


def sanitize_folder_part(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[^\w\-]+", "_", s)
    return s[:80] if s else "NA"


def allowed_file(filename: str) -> bool:
    if not filename or "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in ALLOWED_EXT


def get_next_id() -> int:
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    last_id = 0
    for row in range(ws.max_row, 1, -1):
        val = ws.cell(row=row, column=1).value
        if val is not None and str(val).strip() != "":
            try:
                last_id = int(val)
                break
            except Exception:
                pass
    wb.close()
    return last_id + 1


def append_row(row_values: list):
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.append(row_values)
    wb.save(EXCEL_PATH)
    wb.close()


def _send_report_email(
    report_id: int,
    dt: datetime,
    robots: list[str],
    note: str,
    destinatari: list[str],
    allegati: list[Path],
    *,
    delete_attachments_after_send: bool = False,
    keep_attachments: Optional[list[Path]] = None,
) -> dict:
    """
    Invia email a tutti i destinatari usando il template "REPORT INCIDENTE".
    Dopo l'invio (anche parziale), può eliminare gli allegati passati (tipicamente media temporanei).

    Ritorna un dict con esito per destinatario + file eliminati, utile per debug/log.

    Parametri:
      - delete_attachments_after_send: se True elimina gli allegati dopo l'invio
      - keep_attachments: lista di Path da NON eliminare (es. PDF report)
    """
    results: list[dict] = []
    deleted_files: list[str] = []
    errors: list[str] = []

    if not destinatari:
        return {"ok": True, "results": [], "deleted": [], "errors": []}

    sender = EmailSender()

    template_name = "REPORT INCIDENTE"
    fields = {
        "data": dt.strftime("%d/%m/%Y %H:%M"),
        "robots": ", ".join(robots),
        "note": (note or "").strip(),
    }

    # Allegati reali esistenti
    allegati_paths = [Path(p) for p in (allegati or []) if p]
    attachments = [str(p) for p in allegati_paths if p.exists() and p.is_file()]

    # Invio a tutti
    for dest in destinatari:
        try:
            res = sender.send_template(dest, template_name, fields, attachments)
            results.append(
                {
                    "to": dest,
                    "ok": bool(res.ok),
                    "error": res.error,
                    "log_file": res.log_file,
                    "subject": res.subject,
                    "message_id": res.message_id,
                }
            )
            print(f"[EMAIL] to={dest} ok={res.ok} err={res.error} log={res.log_file}")
        except Exception as e:
            msg = f"[EMAIL] Eccezione invio verso {dest}: {e}"
            print(msg)
            errors.append(msg)
            results.append(
                {
                    "to": dest,
                    "ok": False,
                    "error": str(e),
                    "log_file": None,
                    "subject": None,
                    "message_id": None,
                }
            )

    # Cleanup allegati (tipicamente foto/video) dopo invio
    if delete_attachments_after_send:
        keep_set = set(Path(p) for p in (keep_attachments or []) if p)
        for p in allegati_paths:
            try:
                p = Path(p)
                if p in keep_set:
                    continue
                if p.exists() and p.is_file():
                    p.unlink()
                    deleted_files.append(str(p))
            except Exception as e:
                msg = f"[CLEANUP] Impossibile eliminare {p}: {e}"
                print(msg)
                errors.append(msg)

        if deleted_files:
            print(f"[CLEANUP] Eliminati {len(deleted_files)} file allegati temporanei.")

    # ok globale: almeno un invio riuscito e nessuna eccezione critica
    ok_any = any(r.get("ok") for r in results)
    return {"ok": ok_any, "results": results, "deleted": deleted_files, "errors": errors}

app = Flask(
    __name__,
    static_folder="static",
    static_url_path="/Geekplus/static"
)
app.config["SECRET_KEY"] = "CHANGE_ME__report_medicair_secret"
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_LENGTH


@app.get("/geekplus/reportMedicair")
def report_form():
    now = datetime.now()
    dt_local = now.strftime("%Y-%m-%dT%H:%M")  # datetime-local
    return render_template(
        "reportMedicair.html",
        title=APP_TITLE,
        dt_local=dt_local,
        robot_list=ROBOT_LIST,
        zona_list=ZONA_LIST,
        luci_campo1=LUCI_CAMPO1,
        luci_campo2=LUCI_CAMPO2,
    )


@app.post("/geekplus/reportMedicair")
def submit_report():
    ensure_report_assets()

    dt_local = request.form.get("dt_local", "").strip()
    robots = request.form.getlist("robots")
    scaffale = request.form.get("scaffale", "senza scaffale").strip()
    zona = request.form.get("zona", "").strip()
    errore = request.form.get("errore", "").strip()
    descrizione = request.form.get("descrizione", "").strip()
    luci_c1 = request.form.get("luci_c1", "").strip()
    luci_c2 = request.form.get("luci_c2", "").strip()

    cella = request.form.get("cella", "").strip()
    rimosso = request.form.get("rimosso", "no").strip().lower()
    risoluzione = request.form.get("risoluzione", "").strip()

    errors = []
    if not dt_local:
        errors.append("Data e ora sono obbligatori.")
    if not robots:
        errors.append("Seleziona almeno un robot.")
    if not zona:
        errors.append("Zona è obbligatoria.")
    if not luci_c2:
        errors.append("Luci robot è obbligatorio.")
    if not descrizione:
        errors.append("Descrizione è obbligatoria.")

    invalid = [r for r in robots if r not in ROBOT_LIST]
    if invalid:
        errors.append(f"Robot non validi: {', '.join(invalid)}")
    if zona and zona not in ZONA_LIST:
        errors.append("Zona non valida.")
    if luci_c2 and luci_c2 not in LUCI_CAMPO2:
        errors.append("Luci campo 2 non valide.")
    if rimosso not in ("si", "no"):
        rimosso = "no"

    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("report_form"))

    try:
        dt = datetime.strptime(dt_local, "%Y-%m-%dT%H:%M")
    except Exception:
        flash("Formato data/ora non valido.", "error")
        return redirect(url_for("report_form"))

    next_id = get_next_id()

    folder_stamp = dt.strftime("%d-%m-%Y_%H-%M")
    folder_name = f"{next_id}_{folder_stamp}"
    folder_path = REPORT_DIR / folder_name
    folder_path.mkdir(parents=True, exist_ok=True)

    saved_files: list[str] = []
    saved_file_paths: list[Path] = []

    files = request.files.getlist("media")
    for f in files:
        if not f or not f.filename:
            continue
        if not allowed_file(f.filename):
            continue
        safe = secure_filename(f.filename)
        if not safe:
            continue

        dest = folder_path / safe
        i = 1
        while dest.exists():
            dest = folder_path / f"{dest.stem}_{i}{dest.suffix}"
            i += 1

        f.save(dest)
        saved_files.append(dest.name)
        saved_file_paths.append(dest)

    luci_robot = f"{luci_c1}-{luci_c2}".strip()

    row = [
        next_id,
        dt.strftime("%d/%m/%Y"),
        dt.strftime("%H:%M"),
        ", ".join(robots),
        scaffale if scaffale else "senza scaffale",
        cella,
        zona,
        luci_robot,
        errore,
        descrizione,
        rimosso,
        risoluzione
    ]
    append_row(row)

    pdf_report_path: Path | None = None
    try:
        allegati_list_tex = "\n".join([f"\\item {p.name}" for p in saved_file_paths])

        campi_report = {
            "ReportID": str(next_id),
            "DataIncidente": dt.strftime("%d/%m/%Y"),
            "OraIncidente": dt.strftime("%H:%M"),
            "Robot": ", ".join(robots),
            "Scaffale": scaffale if scaffale else "senza scaffale",
            "Cella": cella,
            "Zona": zona,
            "LuciRobot": luci_robot,
            "Errore": errore,
            "Rimosso": rimosso,
            "Risoluzione": risoluzione,
            "Descrizione": descrizione,
            "AllegatiList": allegati_list_tex,
        }

        nome_pdf = f"REPORT_{next_id}_{dt.strftime('%d-%m-%Y_%H-%M')}"
        res = crea_report("modello", campi_report, nome_file=nome_pdf)

        pdf_report_path = folder_path / res.pdf_path.name
        shutil.copy2(res.pdf_path, pdf_report_path)

        saved_files.append(pdf_report_path.name)
        saved_file_paths.append(pdf_report_path)

    except Exception as e:
        print(f"[PDF] Generazione PDF fallita: {e}")

    # Invio email (media + PDF se disponibile)
    try:
        allegati_email = list(saved_file_paths)
        _send_report_email(
            report_id=next_id,
            dt=dt,
            robots=robots,
            note=descrizione,  # <-- PASSO LA NOTA
            destinatari=list(DESTINATARI),
            allegati=allegati_email,
        )
    except Exception as e:
        print(f"[EMAIL] Invio email fallito: {e}")

    cleanup_latex_tmp()

    return render_template(
        "success.html",
        title=APP_TITLE,
        report_id=next_id,
        folder_name=folder_name,
        saved_files=saved_files,
    )


if __name__ == "__main__":
    ensure_report_assets()
    app.run(host="0.0.0.0", port=3570, debug=False)